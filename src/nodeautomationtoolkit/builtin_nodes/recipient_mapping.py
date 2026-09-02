from __future__ import annotations

import csv
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from nodeautomationtoolkit.core.definition import node
from nodeautomationtoolkit.core.table_types import DataTable

_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"

# Друкується Tkinter-генератором у журнал, щоб одразу було видно, що після
# перезапуску завантажено актуальний вихідний модуль, а не старий процес/EXE.
ROUTING_VERSION = "2026-08-21-v8-multi-tck-kpppo"

_ORDER_SIGNER_START_RE = re.compile(
    r"^\s*(?:т\.?\s*в\.?\s*о\.?|тимчасово\s+виконуюч(?:ий|а)?|"
    r"командувач|командир|начальник|заступник\s+командувача)\b",
    re.IGNORECASE | re.UNICODE,
)

_DISCHARGE_BASIS_RE = re.compile(
    r"У\s+ЗАПАС|ДО\s+ЗАПАСУ|У\s+ВІДСТАВКУ|ЗА\s+ПІДПУНКТ",
    re.IGNORECASE | re.UNICODE,
)


def _is_inline_discharge_basis(text: str) -> bool:
    """Чи містить рядок власну підставу звільнення (У ЗАПАС / ЗА ПІДПУНКТОМ...).

    Використовується, щоб не приліплювати до пункту зовнішню підшапку
    розділу («У ЗАПАС ЗА ПІДПУНКТОМ …:»), якщо пункт сам повністю
    формулює свою підставу звільнення у власному тексті (правило 4
    AGENT.md — підстава належить тексту пункту, а не окремій підшапці).
    """
    return bool(_DISCHARGE_BASIS_RE.search(text))


def _cell_column(reference: str) -> int:
    letters = re.match(r"[A-Z]+", reference.upper())
    result = 0
    for char in letters.group(0) if letters else "A":
        result = result * 26 + ord(char) - 64
    return result - 1


def _read_xlsx(path: Path, sheet_name: str = "") -> list[list[str]]:
    """Читає XLSX за допомогою openpyxl (з data_only=True для обчислення значень формул)."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        rows: list[list[str]] = []
        for r in ws.iter_rows(values_only=True):
            row_vals = [str(c).strip() if c is not None else "" for c in r]
            if any(row_vals):
                rows.append(row_vals)
        wb.close()
        if rows:
            return rows
    except Exception:
        pass

    # Резервний прямий XML парсер (якщо openpyxl відсутній)
    with zipfile.ZipFile(path) as package:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in package.namelist():
            root = ET.fromstring(package.read("xl/sharedStrings.xml"))
            for item in root.findall(f"{{{_MAIN}}}si"):
                shared.append("".join(node.text or "" for node in item.iter(f"{{{_MAIN}}}t")))

        workbook = ET.fromstring(package.read("xl/workbook.xml"))
        relationships = ET.fromstring(package.read("xl/_rels/workbook.xml.rels"))
        targets = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships.findall(f"{{{_PKG_REL}}}Relationship")
        }
        sheets = workbook.find(f"{{{_MAIN}}}sheets")
        if sheets is None or not list(sheets):
            return []
        selected = next(
            (item for item in sheets if item.attrib.get("name") == sheet_name),
            list(sheets)[0],
        )
        rel_id = selected.attrib.get(f"{{{_REL}}}id", "")
        target = targets.get(rel_id, "worksheets/sheet1.xml").lstrip("/")
        target = target if target.startswith("xl/") else "xl/" + target
        sheet = ET.fromstring(package.read(target))
        rows: list[list[str]] = []
        for row in sheet.findall(f".//{{{_MAIN}}}row"):
            values: list[str] = []
            for cell in row.findall(f"{{{_MAIN}}}c"):
                column = _cell_column(cell.attrib.get("r", "A1"))
                while len(values) <= column:
                    values.append("")
                kind = cell.attrib.get("t", "")
                if kind == "inlineStr":
                    value = "".join(
                        item.text or "" for item in cell.iter(f"{{{_MAIN}}}t")
                    )
                else:
                    element = cell.find(f"{{{_MAIN}}}v")
                    value = element.text if element is not None and element.text else ""
                    if kind == "s" and value and value.isdigit() and int(value) < len(shared):
                        value = shared[int(value)]
                values[column] = value.strip()
            if any(values):
                rows.append(values)
        return rows


def _read_rows(path: Path, sheet_name: str = "") -> list[list[str]]:
    if path.suffix.casefold() == ".xlsx":
        return _read_xlsx(path, sheet_name)
    if path.suffix.casefold() != ".csv":
        raise ValueError("Таблиця має бути CSV або XLSX")
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = ";" if raw.count(";") >= raw.count(",") else ","
    reader = csv.reader(raw.splitlines(), delimiter=delimiter)
    return [[cell.strip() for cell in row] for row in reader]


@node(
    name="Прочитати таблицю відповідностей",
    category="Наказ",
    description=(
        "Читає CSV/XLSX таблицю за колонками A (Відкрита назва), B (Закрите найменування / в/ч), "
        "C (Скорочення), D (Корпус), E (Кому / Адресат), F (Куди / Дислокація / Місто)."
    ),
    type_id="builtin.order.read_recipient_mapping",
    execution_inputs=("exec",),
    execution_outputs=("then",),
    outputs={"mapping": "Dictionary", "markers": "List", "table": "DataTable", "count": "int"},
)
def read_recipient_mapping(
    path: str = "",
    open_name_column: str = "Відкрите найменування",
    cipher_column: str = "Шифр",
    destination_column: str = "Куди направляється",
    sheet_name: str = "",
) -> dict:
    source = Path(path).expanduser()
    if not source.is_file():
        raise FileNotFoundError(f"Таблицю не знайдено: {path or '(шлях порожній)'}")
    rows = _read_rows(source, sheet_name)
    if not rows:
        raise ValueError("Таблиця порожня")

    # ── Визначаємо індекси колонок або позиційний доступ A, B, C, D, E, F ──────────
    header_index = -1
    col_a_idx, col_b_idx = 0, 1
    col_c_idx, col_d_idx, col_e_idx, col_f_idx = -1, -1, -1, -1
    has_explicit_abbreviation = False

    for index, row in enumerate(rows[:10]):
        row_norm = {cell.strip().casefold(): i for i, cell in enumerate(row) if cell.strip()}
        if any("відкрит" in k or "найменування" in k for k in row_norm) or any("шифр" in k for k in row_norm):
            header_index = index
            for k, i in row_norm.items():
                if "відкрит" in k or ("назва" in k and "закрит" not in k and "скороч" not in k):
                    col_a_idx = i
                elif "шифр" in k or "закрит" in k or "в/ч" in k or "вч" in k:
                    col_b_idx = i
                elif "скороч" in k or "абрев" in k:
                    col_c_idx = i
                    has_explicit_abbreviation = True
                elif "корпус" in k or k == "ак":
                    col_d_idx = i
                elif "кому" in k or "адресат" in k or "отримувач" in k:
                    col_e_idx = i
                elif "куди" in k or "дислокац" in k or "адрес" in k or "місто" in k:
                    col_f_idx = i
            break

    data_rows = rows[header_index + 1 :] if header_index >= 0 else rows

    # Визначаємо глобальну структуру таблиці
    max_cols = max((len(r) for r in data_rows), default=0)
    is_6_col_format = max_cols >= 5 or col_e_idx >= 0 or col_f_idx >= 0 or col_d_idx >= 0

    c_pos = col_c_idx if col_c_idx >= 0 else 2
    d_pos = col_d_idx if col_d_idx >= 0 else 3
    e_pos = col_e_idx if col_e_idx >= 0 else 4
    f_pos = col_f_idx if col_f_idx >= 0 else 5

    mapping: dict[str, dict[str, str]] = {}
    table_rows = []

    for row in data_rows:
        if not any(row):
            continue
        prow = [str(cell or "").strip() for cell in row]
        while len(prow) < 10:
            prow.append("")

        open_name = prow[col_a_idx if col_a_idx >= 0 else 0]
        cipher = prow[col_b_idx if col_b_idx >= 0 else 1]

        if not is_6_col_format and len(row) <= 3 and col_c_idx < 0:
            abbreviation = ""
            corps = ""
            recipient_to = ""
            destination_where = prow[2]
        elif not is_6_col_format and len(row) == 4 and col_c_idx < 0 and col_d_idx < 0:
            abbreviation = prow[2]
            corps = prow[3]
            recipient_to = ""
            destination_where = ""
        else:
            abbreviation = prow[c_pos] if c_pos != f_pos and c_pos != e_pos else ""
            corps = prow[d_pos] if d_pos != f_pos and d_pos != e_pos else ""
            recipient_to = prow[e_pos]
            destination_where = prow[f_pos]

        if not cipher and not abbreviation and not corps and not recipient_to and not destination_where:
            continue
        # Стовпець A — єдине поле пошуку. Інші стовпці містять лише дані,
        # які використовуються після того, як відповідний рядок уже знайдено.
        if not open_name:
            continue

        destination = destination_where if destination_where else (recipient_to if recipient_to else (corps if corps else (cipher or open_name)))

        entry = {
            "open_name": open_name,
            "cipher": cipher or open_name,
            "abbreviation": abbreviation,
            "corps": corps,  # Якщо D порожнє -> "", відправник сама частина!
            "recipient_to": recipient_to,  # Колонка E: Кому (адресат)
            "destination_where": destination_where,  # Колонка F: Куди (дислокація / місто)
            "dislocation": destination_where or recipient_to,
            "destination": destination,
        }

        mapping[open_name] = entry

        table_rows.append((open_name, cipher, abbreviation, corps, recipient_to, destination_where))

    table = DataTable(
        ("Відкрита назва (A)", "Закрите найменування (B)", "Скорочення (C)", "Корпус (D)", "Кому (E)", "Куди (F)"),
        tuple(table_rows),
        "Таблиця відповідностей ВЧ",
    )
    return {
        "mapping": mapping,
        "markers": list(mapping),
        "table": table,
        "count": len(table_rows),
        # Доказ того, який саме файл був прочитаний у цьому запуску. Дані
        # словника ніде не кешуються: кожен виклик знову читає цей файл з диска.
        "source_path": str(source.resolve()),
        "source_modified_ns": source.stat().st_mtime_ns,
        "source_size": source.stat().st_size,
    }


@node(
    name="Перетворити групи на шифри",
    category="Наказ",
    description=(
        "Зіставляє знайдені відкриті найменування з таблицею, об'єднує однакові "
        "шифри та готує дані для пакета DOCX і звіту."
    ),
    type_id="builtin.order.groups_to_ciphers",
    outputs={
        "documents": "Dictionary",
        "report": "DataTable",
        "missing": "List",
        "summary": "str",
    },
)
def groups_to_ciphers(
    groups: dict | None = None,
    counts: dict | None = None,
    mapping: dict | None = None,
) -> dict:
    documents: dict[str, dict] = {}
    report_rows = []
    missing = []
    for open_name, content in (groups or {}).items():
        entry = (mapping or {}).get(open_name)
        count = int((counts or {}).get(open_name, 0))
        if not isinstance(entry, dict) or not str(entry.get("cipher", "")).strip():
            missing.append(str(open_name))
            report_rows.append((open_name, "", "", count, "Немає відповідності"))
            continue
        cipher = str(entry["cipher"]).strip()
        destination = str(entry.get("destination", "")).strip()
        current = documents.setdefault(
            cipher,
            {
                "name": cipher,
                "cipher": cipher,
                "sender": cipher,
                "destination": destination,
                "content": "",
                "count": 0,
                "open_names": [],
            },
        )
        if current["content"] and str(content).strip():
            current["content"] += "\n"
        current["content"] += str(content).strip()
        current["count"] += count
        current["open_names"].append(str(open_name))
        if not current["destination"]:
            current["destination"] = destination
        report_rows.append((open_name, cipher, destination, count, "Готово"))
    report = DataTable(
        ("Відкрите найменування", "Шифр", "Куди направляється", "Знайдено пунктів", "Стан"),
        tuple(report_rows),
        "Результат групування",
    )
    summary = f"Документів: {len(documents)} · без відповідності: {len(missing)}"
    return {"documents": documents, "report": report, "missing": missing, "summary": summary}


def _normalize_unit_name(name: str) -> str:
    """Нормалізує назву ВЧ для порівняння: видаляє зайві пробіли, переводить у нижній регістр."""
    return re.sub(r"\s+", " ", name.strip().casefold())


_MILITARY_TYPO_DICTIONARY = [
    # Описи та оддруки у слові "бригада / бригади"
    (re.compile(r"\bбригд[иаоеемся]?\b", re.IGNORECASE), "бригади"),
    (re.compile(r"\bбригаи\b", re.IGNORECASE), "бригади"),
    (re.compile(r"\bбриади?\b", re.IGNORECASE), "бригади"),
    (re.compile(r"\bбргади?\b", re.IGNORECASE), "бригади"),
    # Описи у слові "механізована / механізованої"
    (re.compile(r"\bмеханизоаної\b", re.IGNORECASE), "механізованої"),
    (re.compile(r"\bмеханізоаної\b", re.IGNORECASE), "механізованої"),
    (re.compile(r"\bмеханизованої\b", re.IGNORECASE), "механізованої"),
    (re.compile(r"\bмеханизована\b", re.IGNORECASE), "механізована"),
    # Описи у слові "військова частина / в/ч"
    (re.compile(r"\bв\s*\\\s*ч\b", re.IGNORECASE), "в/ч"),
    (re.compile(r"\bв\.?\s*ч\.?\b", re.IGNORECASE), "в/ч"),
    (re.compile(r"\bзв[’'`ʻ]?язку\b", re.IGNORECASE), "зв'язку"),
    (re.compile(r"\bсвязку\b", re.IGNORECASE), "зв'язку"),
    # Описи у типі бригад
    (re.compile(r"\bгірсько\s+штурмов\w*\b", re.IGNORECASE), "гірсько-штурмової"),
    (re.compile(r"\bгірськоштурмов\w*\b", re.IGNORECASE), "гірсько-штурмової"),
    (re.compile(r"\bдесантно\s+штурмов\w*\b", re.IGNORECASE), "десантно-штурмової"),
    (re.compile(r"\bдесантноштурмов\w*\b", re.IGNORECASE), "десантно-штурмової"),
    # Описи для центр та рекрутинг
    (re.compile(r"\bцнтр[уомівіаб]?\b", re.IGNORECASE), "центру"),
    (re.compile(r"\bрекрутинг[уомівіаб]?\b", re.IGNORECASE), "рекрутингу"),
]


def _fix_military_typos(text: str) -> str:
    """Нормалізує поширені описки та варіанти військових найменувань перед обробкою."""
    if not text:
        return ""
    res = text
    for pat, repl in _MILITARY_TYPO_DICTIONARY:
        res = pat.sub(repl, res)
    return res


def _stem_ukrainian_word(word: str) -> str:
    """Видаляє закінчення українських відмінків та стійка до оддруків (БРИГДИ, БРИГАДИ, МЕХАНИЗОВАНОЇ)."""
    w = word.casefold().strip()
    w = re.sub(r"ін$", "он", w)
    w = re.sub(r"іна$", "она", w)

    # Спеціальна обробка військових термінів для захисту від оддруків у документах
    if w.startswith("бриг"):
        return "бриг"
    if w.startswith("механ"):
        return "механ"
    if w.startswith("гірськ") or w.startswith("гірск") or w.startswith("гірш"):
        return "гірс"
    if w.startswith("десант"):
        return "десант"
    if w.startswith("артил"):
        return "артил"
    if w.startswith("баталь") or w.startswith("батальон"):
        return "бат"
    if w.startswith("центр") or w.startswith("цнтр"):
        return "центр"
    if w.startswith("рекрутинг") or w.startswith("рекрут"):
        return "рекрут"

    endings = [
        "ий", "ій", "ого", "ому", "им", "ім", "ої", "ою", "их", "ними", "ими",
        "ами", "ям", "ями", "ях", "ах", "ом", "ем", "єм", "ів", "ев", "єв",
        "и", "і", "а", "я", "у", "ю", "е", "є", "о"
    ]
    for end in sorted(endings, key=len, reverse=True):
        if w.endswith(end) and len(w) - len(end) >= 2:
            w = w[:-len(end)]
            break
    return w


# «о»/«е» в останньому складі зникає при відмінюванні: вузол → вузла,
# орел → орла, вітер → вітру. Шаблон, побудований лише з називного відмінка
# (стовпець A), через це не знаходив ЖОДНОЇ відмінкової форми такої назви.
_FLEETING_VOWEL_RE = re.compile(
    r"(?<=[бвгджзклмнпрстфхцчшщ])[ое]([лкнрцтвбмдгжчшщ])$"
)


def _stem_variants(stem: str) -> list[str]:
    """Стем плюс варіант із випадним голосним.

    Додатковий варіант лише РОЗШИРЮЄ пошук: для назв без чергування
    (полк, центр) регулярка нічого не змінює, бо «о» там не в останньому
    складі, а для нежиттєвих варіантів на кшталт «полігн» у тексті просто
    немає збігів.
    """
    reduced = _FLEETING_VOWEL_RE.sub(r"\1", stem)
    if reduced != stem and len(reduced) >= 3:
        return [stem, reduced]
    return [stem]


def _starts_a_new_heading(line: str) -> bool:
    """Чи є рядок САМОСТІЙНОЮ шапкою, а не уламком перенесеної.

    Справжня підшапка починається з ВЕЛИКИХ літер («У ЗАПАС ЗА ПІДПУНКТОМ …»,
    «ПО ОСОБОВОМУ СКЛАДУ:») або з «Відповідно до»/«Згідно з». Уламок шапки,
    розірваної мʼяким переносом, — звичайний текст, який може починатися і з
    лапки («“Про військовий обовʼязок…»), тому перевірка «перша літера мала»
    його не ловила: шапка ставала двома сегментами, обидва вказували на ОДИН
    абзац Word — і у витягу вона друкувалася ДВІЧІ.
    """
    clean = str(line or "").strip()
    if clean.startswith("§"):
        return True
    low = clean.casefold()
    if low.startswith("відповідно до") or low.startswith("згідно з"):
        return True
    words = [w for w in re.split(r"[^A-Za-zА-Яа-яІіЇїЄєҐґ']+", clean) if w][:2]
    if not words:
        return False
    first = words[0]
    if len(first) >= 2 and first.isupper():
        return True
    if len(first) == 1 and first.isupper() and len(words) > 1:
        second = words[1]
        return len(second) >= 2 and second.isupper()
    return False


def _build_unit_fuzzy_pattern(open_name: str) -> re.Pattern:
    """
    Будує нечіткий регулярний вираз для пошуку назви військової частини у будь-яких відмінках.
    Враховує відмінювання (полк, полку, полком, бригади, бригадою), апострофи та розширену відстань між словами.
    """
    STOP_WORDS = {
        "та", "і", "й", "з", "зі", "із", "на", "в", "у", "до", "від", "по", "при", "за",
        "the", "of", "and", "а", "ім", "імені",
        "командний", "командного", "командним", "командному", "пункт", "пункту", "пунктом", "пункті",
        "передовий", "передового", "передовим", "передовому", "запасний", "запасного", "запасним", "запасному",
        "зкп", "гкп", "ппу", "кп", "управління",
    }

    clean_name = re.sub(r"[^\w\s\d]", "'", open_name)
    tokens = [t for t in re.split(r"[^\w\d]+", clean_name) if t]

    anchors: list[str] = []
    prefix_anchors: list[str] = []
    for token in tokens:
        if not token:
            continue
        if re.match(r"^\d+$", token):
            # Номер частини має збігатися повністю: «8» не є «18» або «80».
            numeric_anchor = rf"(?<!\d){re.escape(token)}(?!\d)"
            anchors.append(numeric_anchor)
            prefix_anchors.append(numeric_anchor)
        elif token.casefold() in STOP_WORDS:
            continue
        elif len(token) >= 2:
            stem = _stem_ukrainian_word(token)
            escaped_variants = [
                re.escape(variant).replace("\\'", r"[^\w\s]?").replace("'", r"[^\w\s]?")
                for variant in _stem_variants(stem)
            ]
            if len(escaped_variants) > 1:
                pattern_part = "(?:" + "|".join(escaped_variants) + r")\w*"
            else:
                pattern_part = escaped_variants[0] + r"\w*"
            anchors.append(pattern_part)

            # Запасний загальний пошук за стабільним початком КОЖНОГО
            # значущого слова. Він не залежить від переліку відмінкових
            # закінчень: для довгих слів достатньо перших 4 літер. Якщо в
            # основі є чергування (вузол/вузла, загін/загону), беремо 3.
            prefix_base = token.casefold()
            variants = _stem_variants(stem)
            has_early_alternation = (
                len(variants) > 1
                and len({variant[:4] for variant in variants}) > 1
            ) or bool(re.search(r"ін(?:а|у|ом|і)?$", prefix_base))
            if len(prefix_base) <= 4:
                prefix_length = len(prefix_base)
            elif has_early_alternation:
                prefix_length = 3
            else:
                # Префікс росте з довжиною слова, але не довший за 6.
                # Фіксовані 4 зливали слова зі спільним коренем:
                # «радіотехнічний» давав «раді», що ловило «РАДІОРЕЛЕЙНОГО»,
                # і витяг ішов ще й на «99 окремий радіотехнічний батальйон»,
                # хоча в наказі був «99 окремий полк звʼязку».
                # Довший за 6 теж не можна: «окремий» дало б «окреми», яке не
                # ловить «окремого», а «відновлювальний» перестало б ловити
                # скорочену форму «відновного». Обидві пари розходяться саме
                # на 6-му символі, тому 6 — стеля, а не бажане значення.
                prefix_length = max(4, min(6, len(prefix_base) - 3))
            stable_prefix = prefix_base[:prefix_length]
            prefix_anchors.append(re.escape(stable_prefix) + r"\w*")

    if not anchors:
        return re.compile(re.escape(open_name), re.IGNORECASE)

    # Проміжок між словами назви навмисно широкий — між ними стоять почесні
    # найменування («ордена Богдана Хмельницького», «імені гетьмана …»).
    # Але він НЕ МОЖЕ містити номер іншої частини: інакше назва «зшивалася»
    # з уламків двох сусідніх частин у переліку, напр. шаблон
    # «158 окрема бригада підтримки» хибно збігався з текстом
    # «158 окремого батальйону зв'язку та 47 окремої бригади підтримки».
    # Проміжок так само НЕ МОЖЕ перетнути тире-роздільник « – », яким у пункті
    # відділені «звідки» і «КУДИ». Інакше збіг починався на «…Тестівської
    # області» й тягнувся аж до «…СОЦІАЛЬНОЇ ПІДТРИМКИ» у ВЕЛИКІЙ половині
    # пункту: 161 символ разом із «– НАЧАЛЬНИКОМ ГРУПИ …» замінювався одним
    # шифром, і 139 символів наказу ЗНИКАЛИ. Тире всередині слова
    # («гірсько-штурмової») не заважає: перевіряється тире з пробілами обабіч.
    gap = r"(?:(?!\b\d{1,4}\b)(?!\s[-–—]\s)[\s\S]){0,180}?"
    full_stem_pattern = gap.join(anchors)
    stable_prefix_pattern = gap.join(prefix_anchors)
    alternatives = [full_stem_pattern, stable_prefix_pattern]

    # Для командного пункту ППО стабільним ідентифікатором є номер та ядро
    # типу частини. У колонці A після нього можуть стояти додаткові слова
    # підпорядкування, яких у конкретному пункті наказу немає. Вимагати всі
    # слова такого довгого рядка означало пропускати явне
    # «10 командного пункту протиповітряної оборони».
    name_low = open_name.casefold()
    number_match = re.search(r"(?<!\d)(\d{1,4})(?!\d)", name_low)
    is_air_defense_command_post = (
        number_match is not None
        and (
            (
                "команд" in name_low
                and "пункт" in name_low
                and "протиповітр" in name_low
                and "оборон" in name_low
            )
            or "кпппо" in re.sub(r"[^а-яіїєґa-z0-9]", "", name_low)
        )
    )
    if is_air_defense_command_post:
        number_anchor = rf"(?<!\d){re.escape(number_match.group(1))}(?!\d)"
        core_gap = r"(?:(?!\b\d{1,4}\b)(?!\s[-–—]\s)[\s\S]){0,120}?"
        expanded_core = core_gap.join(
            (number_anchor, r"команд\w*", r"пункт\w*", r"протиповітр\w*", r"оборон\w*")
        )
        acronym_core = number_anchor + r"[\s\-]*кпппо\w*"
        alternatives.extend((expanded_core, acronym_core))

    pattern_str = "(?:" + "|".join(alternatives) + ")"
    return re.compile(pattern_str, re.IGNORECASE | re.UNICODE | re.DOTALL)


_CORPS_RE = re.compile(
    r"\b(\d{1,3})[-\s]*(?:-?й|-?го|-?му|-?м)?\s*(?:армійськ\w*\s+корпус\w*|АК)\b",
    re.IGNORECASE | re.UNICODE,
)


def _extract_army_corps(text: str) -> str | None:
    """
    Знаходить армійський корпус у тексті (наприклад: '10-го армійського корпусу', '3 АК').
    Повертає нормалізовану назву: '10 армійський корпус'.
    """
    match = _CORPS_RE.search(text)
    if match:
        num = match.group(1)
        return f"{num} армійський корпус"
    return None


_RECRUITING_CENTER_9_RE = re.compile(
    r"\b9[-\s]*(?:-?й|-?го|-?му|-?м)?\s*(?:центр\w*|цнтр\w*)\s+рекрутинг\w*\b",
    re.IGNORECASE | re.UNICODE,
)


# Слово «територіальний» у наказі часто пропускають: «ХМЕЛЬНИЦЬКОГО ОБЛАСНОГО
# ЦЕНТРУ КОМПЛЕКТУВАННЯ ТА СОЦІАЛЬНОЇ ПІДТРИМКИ» — без нього. Через це пункт
# не знаходив адресата ЗОВСІМ і потрапляв у «Пропущені». Ядро назви —
# «центр комплектування», воно й лишається обовʼязковим.
_TCK_KEYWORDS_RE = re.compile(
    r"(?:(?:територіальн\w*\s+)?центр\w*\s+комплектування\w*|ТЦК\w*|РТЦК\w*|МТЦК\w*|ОТЦК\w*)",
    re.IGNORECASE | re.UNICODE,
)

_TCK_OBLAST_EXPLICIT_RE = re.compile(
    r"\b([А-ЯІЇЄа-яіїє'ʼ-]+?(?:ськ|цьк|зьк)\w*)\s+област",
    re.IGNORECASE | re.UNICODE,
)

_TCK_EXPLICIT_CITY_RE = re.compile(
    r"\b(?:у|в)\s+(?:місті|м\.)\s*([А-ЯІЇЄа-яіїє'ʼ-]+)",
    re.IGNORECASE | re.UNICODE,
)

_TCK_REGION_BEFORE_RE = re.compile(
    r"\b([А-ЯІЇЄа-яіїє'ʼ-]+?(?:ськ|цьк|зьк)\w*)"
    r"\s+(?:районн\w*|міськ\w*|обласн\w*|)?\s*"
    r"(?:територіальн\w*\s+центр\w*\s+комплектування\w*|[РМОО]?ТЦК\w*)",
    re.IGNORECASE | re.UNICODE,
)

_RAYON_TO_OBLAST_MAP = {
    "броварськ": "Київський",
    "білоцерківськ": "Київський",
    "бориспільськ": "Київський",
    "вишгородськ": "Київський",
    "фастівськ": "Київський",
    "обухівськ": "Київський",
    "бучанськ": "Київський",
    "ковельськ": "Волинський",
    "володимирськ": "Волинський",
    "камінь-каширськ": "Волинський",
    "дрогобицьк": "Львівський",
    "стрийськ": "Львівський",
    "самбірськ": "Львівський",
    "червоноградськ": "Львівський",
    "золочівськ": "Львівський",
    "яворівськ": "Львівський",
    "криворізьк": "Дніпропетровський",
    "нікопольськ": "Дніпропетровський",
    "павлоградськ": "Дніпропетровський",
    "новомосковськ": "Дніпропетровський",
    "ізмаїльськ": "Одеський",
    "болградськ": "Одеський",
    "білгород-дністровськ": "Одеський",
    "ізюмськ": "Харківський",
    "куп'янськ": "Харківський",
    "чугуївськ": "Харківський",
    "кременчуцьк": "Полтавський",
    "лубенськ": "Полтавський",
    "миргородськ": "Полтавський",
    "бердичівськ": "Житомирський",
    "коростенськ": "Житомирський",
    "новоград-волинськ": "Житомирський",
    "звагельськ": "Житомирський",
    "ужгородськ": "Закарпатський",
    "мукачівськ": "Закарпатський",
    "уманськ": "Черкаський",
    "прилуцьк": "Чернігівський",
    "ніжинськ": "Чернігівський",
    "конотопськ": "Сумський",
    "шосткинськ": "Сумський",
    "сарненськ": "Рівненський",
    "дубенськ": "Рівненський",
    "хмільницьк": "Вінницький",
    "жмеринськ": "Вінницький",
}

_UKRAINE_OBLAST_STEMS = {
    "вінниц": "Вінницький",
    "волин": "Волинський",
    "дніпро": "Дніпропетровський",
    "дніпропетровськ": "Дніпропетровський",
    "донец": "Донецький",
    "житомир": "Житомирський",
    "закарпат": "Закарпатський",
    "запоріж": "Запорізький",
    "запорізьк": "Запорізький",
    "івано-франків": "Івано-Франківський",
    "івано-франківськ": "Івано-Франківський",
    "франків": "Івано-Франківський",
    "франківськ": "Івано-Франківський",
    "київ": "Київський",
    "київськ": "Київський",
    "києв": "Київський",
    "кіровоград": "Кіровоградський",
    "луган": "Луганський",
    "львів": "Львівський",
    "львов": "Львівський",
    "миколаїв": "Миколаївський",
    "миколаєв": "Миколаївський",
    "одес": "Одеський",
    "одеськ": "Одеський",
    "полтав": "Полтавський",
    "рівнен": "Рівненський",
    "сум": "Сумський",
    "тернопіл": "Тернопільський",
    "тернопол": "Тернопільський",
    "харків": "Харківський",
    "харков": "Харківський",
    "херсон": "Херсонський",
    "хмельниц": "Хмельницький",
    "черкас": "Черкаський",
    "чернівец": "Чернівецький",
    "чернігів": "Чернігівський",
    "чернігов": "Чернігівський",
}


def _normalize_region_to_nominative(region_raw: str) -> str:
    """Нормалізує прикметник регіону до називного відмінку чоловічого роду (напр. Камінь-Каширського → Камінь-Каширський, Івано-Франківського → Івано-Франківський)."""
    raw = (region_raw or "").strip()
    if raw.lower().startswith("івано-франків"):
        return "Івано-Франківський"
    parts = raw.split("-")
    norm_parts = []
    for part in parts:
        p = part.strip()
        if p.lower() in ("івано", "камінь", "кам'янець", "рава", "кривий", "дніпро"):
            norm_parts.append(p[0].upper() + p[1:] if p else "")
            continue
        p = re.sub(r"(?:ського|ської|ському|ським|ською)$", "ський", p, flags=re.IGNORECASE)
        p = re.sub(r"(?:цького|цької|цькому|цьким|цькою)$", "цький", p, flags=re.IGNORECASE)
        p = re.sub(r"(?:зького|зької|зькому|зьким|зькою)$", "зький", p, flags=re.IGNORECASE)
        if not (p.lower().endswith("ський") or p.lower().endswith("цький") or p.lower().endswith("зький") or p.lower().endswith("ін")):
            p += "ський"
        norm_parts.append(p[0].upper() + p[1:] if p else "")
    return "-".join(norm_parts)


_OBLAST_TO_CITY_MAP = {
    "вінниц": "м. Вінниця",
    "волин": "м. Луцьк",
    "дніпро": "м. Дніпро",
    "дніпропетровськ": "м. Дніпро",
    "донец": "м. Краматорськ",
    "житомир": "м. Житомир",
    "закарпат": "м. Ужгород",
    "запоріж": "м. Запоріжжя",
    "івано-франків": "м. Івано-Франківськ",
    "київ": "м. Київ",
    "кіровоград": "м. Кропивницький",
    "луган": "м. Сєвєродонецьк",
    "львів": "м. Львів",
    "миколаїв": "м. Миколаїв",
    "одес": "м. Одеса",
    "полтав": "м. Полтава",
    "рівнен": "м. Рівне",
    "сум": "м. Суми",
    "тернопіл": "м. Тернопіль",
    "харків": "м. Харків",
    "херсон": "м. Херсон",
    "хмельниц": "м. Хмельницький",
    "черкас": "м. Черкаси",
    "чернівец": "м. Чернівці",
    "чернігів": "м. Чернігів",
}


def _find_entry_in_mapping(norm_code: str, open_name: str, mapping_dict: dict) -> dict | None:
    """Знаходить рядок таблиці виключно за його пошуковою назвою зі стовпця A."""
    if not mapping_dict:
        return None
    candidates = {
        _normalize_unit_name(value)
        for value in (norm_code, open_name)
        if str(value or "").strip()
    }
    for k, val in mapping_dict.items():
        if not isinstance(val, dict):
            continue
        column_a = str(val.get("open_name") or k).strip()
        if _normalize_unit_name(column_a) in candidates:
            return val

    # Районний/міський ТЦК зводиться до обласного, але сам рядок однаково
    # вибирається лише за назвою зі стовпця A.
    norm_low = (norm_code or "").casefold()
    open_low = (open_name or "").casefold()
    if "тцк" in norm_low or "центр" in norm_low or "отцк" in norm_low or "тцк" in open_low:
        for k, val in mapping_dict.items():
            if not isinstance(val, dict):
                continue
            column_a_low = str(val.get("open_name") or k).casefold()
            for stem in _UKRAINE_OBLAST_STEMS:
                if (stem in norm_low or stem in open_low) and (
                    stem in column_a_low
                ):
                    return val
    return None




def _extract_tck_sender(text: str) -> str | None:
    """Розпізнає ТЦК та СП, а також 9 Центр рекрутингу (тимчасове правило розсилання на ТЦК)."""
    is_tck = bool(_TCK_KEYWORDS_RE.search(text))
    is_recruiting_9 = bool(_RECRUITING_CENTER_9_RE.search(text))

    if not (is_tck or is_recruiting_9):
        return None

    # Пріоритет 1: Якщо у тексті прямо вказано назву області (напр: Київської області)
    oblast_match = _TCK_OBLAST_EXPLICIT_RE.search(text)
    if oblast_match:
        reg_nom = _normalize_region_to_nominative(oblast_match.group(1))
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if reg_nom.lower().startswith(stem):
                return f"{obl_name} ОТЦК та СП"
        return f"{reg_nom} ОТЦК та СП"

    # Пріоритет 2: явне місце розташування «у місті Львові» переважає
    # складену районну назву. Наприклад, «Галицько-Франківського об'єднаного
    # районного у місті Львові ТЦК» належить Львівському ОТЦК, а слово
    # «Франківського» не означає Івано-Франківську область.
    city_match = _TCK_EXPLICIT_CITY_RE.search(text)
    if city_match:
        city_low = city_match.group(1).casefold()
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if city_low.startswith(stem):
                return f"{obl_name} ОТЦК та СП"

    # Пріоритет 3: Шукаємо прикметник БЕЗПОСЕРЕДНЬО перед ТЦК (напр: Броварського РТЦК)
    region_match = _TCK_REGION_BEFORE_RE.search(text)
    if region_match:
        reg_nom = _normalize_region_to_nominative(region_match.group(1))
        low = reg_nom.lower()
        # Перевіряємо чи це область
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if low.startswith(stem):
                return f"{obl_name} ОТЦК та СП"
        # Перевіряємо чи це район відомої області
        for r_stem, obl_name in _RAYON_TO_OBLAST_MAP.items():
            if low.startswith(r_stem):
                return f"{obl_name} ОТЦК та СП"
        # Перевіряємо чи в тексті є згадка обласного центру чи області (напр. м. Львова, м. Одеси)
        low_full = text.lower()
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if stem in low_full:
                return f"{obl_name} ОТЦК та СП"
        return f"{reg_nom} ОТЦК та СП"

    low_full = text.lower()
    for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
        if stem in low_full:
            return f"{obl_name} ОТЦК та СП"

    return "Обласний ТЦК та СП"


def _extract_tck_senders(text: str) -> list[str]:
    """Повертає всі обласні ТЦК, явно названі в одному пункті.

    Пункт про скасування/переміщення може одночасно згадувати ТЦК двох
    областей. Старий одиничний extractor повертав лише перший і втрачав
    другого адресата.
    """
    if not (_TCK_KEYWORDS_RE.search(text or "") or _RECRUITING_CENTER_9_RE.search(text or "")):
        return []

    senders: list[str] = []

    def add_oblast(raw_region: str) -> None:
        reg_nom = _normalize_region_to_nominative(raw_region)
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if reg_nom.casefold().startswith(stem):
                sender = f"{obl_name} ОТЦК та СП"
                if sender not in senders:
                    senders.append(sender)
                return

    # Явно названі області мають найвищий пріоритет і можуть бути різними.
    for match in _TCK_OBLAST_EXPLICIT_RE.finditer(text or ""):
        add_oblast(match.group(1))
    if senders:
        return senders

    # Якщо областей немає, збираємо всі явні місця «у місті ...».
    for match in _TCK_EXPLICIT_CITY_RE.finditer(text or ""):
        city_low = match.group(1).casefold()
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if city_low.startswith(stem):
                sender = f"{obl_name} ОТЦК та СП"
                if sender not in senders:
                    senders.append(sender)
                break
    if senders:
        return senders

    fallback = _extract_tck_sender(text)
    return [fallback] if fallback else []

def _extract_tck_region_hints(text: str) -> list[str]:
    """Витягує назви областей з тексту ТЦК."""
    hints = []
    oblast_match = _TCK_OBLAST_EXPLICIT_RE.search(text or "")
    if oblast_match:
        reg_nom = _normalize_region_to_nominative(oblast_match.group(1))
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if reg_nom.lower().startswith(stem):
                hints.append(obl_name)
        if not hints:
            hints.append(reg_nom)
    city_match = _TCK_EXPLICIT_CITY_RE.search(text or "")
    if city_match:
        city_low = city_match.group(1).casefold()
        for stem, obl_name in _UKRAINE_OBLAST_STEMS.items():
            if city_low.startswith(stem) and obl_name not in hints:
                hints.append(obl_name)
    return hints


def _extract_corps_abbr(corps_str: str) -> str:
    """Витягує скорочену назву корпусу (напр. '25 армійський корпус' -> '25АК')."""
    clean = str(corps_str).strip()
    match = re.search(r"\b(\d{1,3})\s*(?:армійськ\w*\s+корпус\w*|АК)\b", clean, re.IGNORECASE)
    if match:
        return f"{match.group(1)}АК"
    return clean


def _normalize_key(raw_k: str, canonical_map: dict[str, str] | None = None) -> str:
    k = str(raw_k).strip()
    cmap = canonical_map or {}
    if k in cmap:
        return cmap[k]
    short_k = _short_closed_code(k)
    if short_k in cmap:
        return cmap[short_k]
    c_abbr = _extract_corps_abbr(k)
    if c_abbr in cmap:
        return cmap[c_abbr]
    return k


def _auto_abbreviate_unit_name(open_name: str) -> str:
    """Автоматично генерує скорочення ВЧ з повної назви (напр. '15 окрема механізована бригада' -> '15омбр')."""
    clean = str(open_name).strip()
    match_num = re.search(r"\b(\d{1,4})\b", clean)
    num_str = match_num.group(1) if match_num else ""

    low = clean.lower()
    type_abbr = ""
    if "механізован" in low or "омбр" in low:
        type_abbr = "омбр"
    elif "танков" in low or "отбр" in low:
        type_abbr = "отбр"
    elif "десантно-штурмов" in low or "одшбр" in low:
        type_abbr = "одшбр"
    elif "гірсько-штурмов" in low or "огшбр" in low:
        type_abbr = "огшбр"
    elif "артилерійськ" in low or "оабр" in low:
        type_abbr = "оабр"
    elif "полк зв" in low or "зв’язку" in low or "зв'язку" in low or "опз" in low:
        type_abbr = "опз"
    elif "розвідувальн" in low or "орб" in low:
        type_abbr = "орб"
    elif "інженерн" in low or "оібр" in low:
        type_abbr = "оібр"
    elif "автомобільн" in low or "оаб" in low:
        type_abbr = "оаб"
    elif "центр рекрутинг" in low or "цр" in low:
        type_abbr = "ЦР"
    elif "батальйон" in low:
        type_abbr = "ОБ"

    if num_str and type_abbr:
        return f"{num_str}{type_abbr}"
    return clean


def _norm_corps_token(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _find_corps_entry(corps_col_val: str, corps_abbr_val: str, mapping_dict: dict) -> dict | None:
    """Знаходить РЯДОК корпусу, на який посилається стовпець D частини.

    Це розвʼязування внутрішнього посилання таблиці (D → рядок), а не пошук
    у тексті наказу, тому тут дозволено дивитися і на стовпець C. Пошук
    адресата в тексті, як і раніше, йде виключно за стовпцем A.
    """
    entry = mapping_dict.get(corps_col_val) or mapping_dict.get(corps_abbr_val)
    if isinstance(entry, dict) and entry.get("cipher"):
        return entry

    # Стовпець D часто містить СКОРОЧЕННЯ корпусу («ОК Захід»), яке збігається
    # зі стовпцем C його рядка, а не зі стовпцем A («оперативне командування
    # «Захід»») і не з шаблоном «N АК». Без цього рядок корпусу не знаходився:
    # ключ будувався без шифру, і на ОДИН корпус виходило ДВА витяги —
    # «ОК Захід» (через підпорядковану частину) і «ОК Захід А0777» (від
    # прямого збігу), причому в першому «Кому»/«Куди» лишалися від частини.
    wanted = {_norm_corps_token(corps_col_val), _norm_corps_token(corps_abbr_val)} - {""}
    if wanted:
        for _name, _val in mapping_dict.items():
            if not isinstance(_val, dict) or not _val.get("cipher"):
                continue
            abbreviation = _norm_corps_token(_val.get("abbreviation") or _val.get("abbr"))
            if abbreviation and abbreviation in wanted:
                return _val

    for _name, _val in mapping_dict.items():
        if not isinstance(_val, dict):
            continue
        column_a = str(_val.get("open_name") or _name).strip()
        if _extract_corps_abbr(column_a) == corps_abbr_val and _val.get("cipher"):
            return _val
    return None


def _build_sender_key(closed_code: str, abbreviation: str, corps_col: str, corps_resolved_cipher: dict, mapping_dict: dict) -> tuple[str, str]:
    """
    Будує точний sender_key для ВЧ/Корпусу, повністю зберігаючи ТЕКСТ З КЛІТИНКИ СКОРОЧЕНОЇ НАЗВИ (колонка C).
    Усуває дублювання шифрів та усуває дублі у підсумковому виводі.
    """
    short_cipher = _short_closed_code(closed_code)

    if corps_col:
        corps_abbr = _extract_corps_abbr(corps_col)
        corps_cipher = corps_resolved_cipher.get(corps_abbr, short_cipher)
        corps_entry = _find_corps_entry(corps_col, corps_abbr, mapping_dict)
        if corps_entry:
            # Ключ корпусу будуємо ТИМИ САМИМИ правилами, що й для самого
            # корпусу: інакше корпус із порожньою колонкою C отримував два
            # різні ключі («14АК А5555» через частину і «А5555» напряму),
            # і на один корпус створювалося два витяги.
            sender_key, _ = _build_sender_key(
                str(corps_entry.get("cipher") or corps_cipher),
                str(corps_entry.get("abbreviation") or corps_entry.get("abbr") or "").strip(),
                "",
                corps_resolved_cipher,
                mapping_dict,
            )
        else:
            sender_key = f"{corps_abbr} {corps_cipher}".strip() if corps_cipher else corps_abbr
        return sender_key, corps_abbr

    abbr = str(abbreviation).strip()
    if not abbr or abbr.isdigit():
        return short_cipher, short_cipher

    # Для ТЦК та СП — повертаємо скорочення як є
    if "ТЦК" in abbr.upper() or "ОТЦК" in abbr.upper() or "РТЦК" in abbr.upper() or "МТЦК" in abbr.upper():
        return abbr, abbr

    # Якщо у клітинці скорочення (колонка C) ВЖЕ є шифр (напр. '24ОМБр А2424') — копіюємо ВСЕ з клітинки
    if re.search(r"\bА\s*\d{4}\b", abbr, re.IGNORECASE) or (short_cipher and short_cipher.casefold() in abbr.casefold()):
        return abbr, abbr

    # Інакше додаємо шифр до скорочення з колонки C: '24ОМБр А2424'
    return f"{abbr} {short_cipher}", abbr


def _short_closed_code(code: str, abbreviation: str = "") -> str:
    """Формує коротке закрите найменування частини або ТЦК (усуває дублювання довгих назв ТЦК)."""
    clean_code = str(code).strip()
    if clean_code.lower().startswith("в/ч "):
        cipher = clean_code[4:].strip()
    elif clean_code.lower().startswith("в/ч"):
        cipher = clean_code[3:].strip()
    else:
        cipher = clean_code

    abbr = str(abbreviation).strip()

    # Для ТЦК та СП джерелом є виключно таблиця: точний текст колонки C,
    # а якщо вона порожня — точний текст колонки B. Не нормалізуємо й не
    # конструюємо назву з окремо знайденого прикметника.
    if "ТЦК" in abbr.upper() or "ТЦК" in cipher.upper() or "ТЕРИТОРІАЛЬН" in cipher.upper() or "ЦЕНТЕР" in cipher.upper() or "ЦЕНТР" in cipher.upper():
        return abbr or cipher

    if abbr and abbr != cipher and abbr.casefold() not in cipher.casefold():
        return f"{abbr} {cipher}"
    return cipher


def _format_item_numbers_range(labels: list[str]) -> str:
    """Форматує перелік пунктів з використанням дефісів для послідовних діапазонів (напр. 1,3,4-7,9,10,11-25).
    Для ієрархічних номерів (1.1, 1.2) повертає точний перелік через кому."""
    has_hierarchical = any(re.search(r"\b\d+\.\d+\b", str(lbl)) for lbl in labels)
    if has_hierarchical:
        items = []
        for label in labels:
            clean = str(label).strip()
            m = re.search(r"\b\d+(?:\.\d+)+\b|\b\d+\b", clean)
            if m:
                items.append(m.group(0))
            elif clean:
                items.append(clean)
        return ",".join(items) if items else "-"

    nums = set()
    non_nums = []

    for label in labels:
        clean = str(label).strip()
        matches = re.findall(r"\b\d+\b", clean)
        if matches:
            nums.add(int(matches[0]))
        elif clean:
            non_nums.append(clean)

    sorted_nums = sorted(nums)
    if not sorted_nums:
        return ",".join(non_nums) if non_nums else "-"

    ranges = []
    start = sorted_nums[0]
    end = sorted_nums[0]

    for n in sorted_nums[1:]:
        if n == end + 1:
            end = n
        else:
            if end > start + 1:
                ranges.append(f"{start}-{end}")
            elif end == start + 1:
                ranges.append(f"{start},{end}")
            else:
                ranges.append(f"{start}")
            start = n
            end = n

    if end > start + 1:
        ranges.append(f"{start}-{end}")
    elif end == start + 1:
        ranges.append(f"{start},{end}")
    else:
        ranges.append(f"{start}")

    result_str = ",".join(ranges)
    if non_nums:
        result_str += "," + ",".join(non_nums)
    return result_str



def _get_item_main_text(lines: list[str]) -> str:
    """
    Витягує виключно розпорядчий текст пункту наказу.
    Повністю ігнорує:
    1. Рядки підстав ('Підстава: ...').
    2. Біографічні та облікові рядки (р.н., освіта, ІПН/РНОКПП, ВОС, 'у ЗС із...',
       'призваний...', 'перебуває на обліку...', 'підлягає направленню на військовий облік...').
    """
    hard_bio_markers = (
        "народивс",
        "народил",
        "р.н.",
        "р. н.",
        "року народження",
        "рнокпп",
        "іпн",
        "освіта",
        "вос-",
        "вос ",
        "у зс",
        "у зсу",
        "призваний",
        "призвана",
    )
    accounting_markers = (
        "військовий облік",
        "військовому обліку",
        "військового обліку",
        "підлягає",
        "підлягають",
        "направленню на",
        "перебуває на",
        "перебувають на",
    )
    selected_lines = []
    # Біографічний блок — це ХВІСТ пункту: після нього розпорядчого тексту вже
    # немає. Без цього мʼякий перенос усередині маркера ламав фільтр: рядок
    # «Підлягає направленню на військовий ⏎ облік до …ТЦК та СП» розпадався на
    # два, у першому маркер був, а другий («облік до …») не збігався з жодним —
    # і ТЦК ставав адресатом, тобто витяг ішов не туди.
    tail_started = False
    for line in lines:
        clean = line.strip()
        low = clean.lower()
        if not clean:
            continue
        if tail_started:
            continue
        if clean.startswith("Підстава") or clean.startswith("підстава"):
            continue
        # У частині наказів розпорядчий текст і біографічні відомості стоять
        # в одному абзаці. Не відкидаємо весь абзац: залишаємо початок із
        # адресатом зі стовпця A та відрізаємо лише біографічний хвіст.
        cutoff_positions = [low.find(marker) for marker in hard_bio_markers if marker in low]
        if cutoff_positions:
            main_prefix = clean[: min(cutoff_positions)].rstrip(" ,;:–—-")
            if main_prefix:
                selected_lines.append(main_prefix)
            tail_started = True
            continue
        accounting_positions = [low.find(marker) for marker in accounting_markers if marker in low]
        if accounting_positions:
            accounting_start = min(accounting_positions)
            prefix = clean[:accounting_start]
            has_item_label = bool(
                re.match(r"^\s*\d{1,3}(?:\.\d{1,3})*[\.\)]\s*", clean)
            )
            prefix_without_label = re.sub(
                r"^\s*\d{1,3}(?:\.\d{1,3})*[\.\)]\s*",
                "",
                prefix,
            ).strip(" ,;:–—-")
            # Окремий абзац військового обліку є біографічним. Якщо ж перед
            # ним у тому самому абзаці є розпорядчий текст або це самостійний
            # пронумерований пункт, не втрачаємо названий там маршрут ТЦК.
            if not prefix_without_label and not has_item_label:
                tail_started = True
                continue
        selected_lines.append(line)
    return "\n".join(selected_lines)


@node(
    name="Картування та пошук військових частин",
    category="Наказ",
    description=(
        "Шукає відкриті найменування військових частин у тексті наказу, замінює їх на закриті "
        "найменування (шифри / в/ч), зберігає шапки, преамбули (§) та повні блоки пунктів з Підставами. "
        "Підтримує нечіткий відмінковий пошук: знаходить 'бригади' замість 'бригада', "
        "'механізованої' замість 'механізована' тощо."
    ),
    type_id="builtin.order.map_military_units",
    outputs={
        "processed_text": "str",
        "units_table": "DataTable",
        "units_list": "List",
        "unit_paragraphs": "Dictionary",
        "match_report": "DataTable",
        "summary": "str",
    },
)
def map_military_units(
    text: str = "",
    mapping: dict | None = None,
    default_prefix: str = "військова частина ",
    fuzzy_match: bool = True,
) -> dict:
    if not text.strip():
        return {
            "processed_text": "",
            "units_table": DataTable(("Закрита ВЧ", "Відкрита назва", "Пункти наказу", "Згадок"), ()),
            "units_list": [],
            "unit_paragraphs": {},
            "match_report": DataTable(("Шаблон пошуку", "Знайдений збіг", "Шифр", "Кількість"), ()),
            "summary": "Порожній текст наказу",
        }

    text = _fix_military_typos(text)
    mapping_raw = mapping or {}
    mapping_dict = {}
    for k, v in mapping_raw.items():
        clean_k = _fix_military_typos(str(k))
        if isinstance(v, dict):
            clean_v = dict(v)
            if "open_name" in clean_v:
                clean_v["open_name"] = _fix_military_typos(str(clean_v["open_name"]))
            mapping_dict[clean_k] = clean_v
        else:
            mapping_dict[clean_k] = v

    lines = [line.rstrip() for line in text.splitlines()]

    # ── Визначаємо шапку наказу ────────────────────────────────────────────────
    header_lines = []
    content_start_idx = len(lines)
    for idx, line in enumerate(lines):
        clean = line.strip()
        clean_search = re.sub(r"\s+", " ", clean)
        if (
            clean_search.startswith("§")
            or re.match(r"^\d+[\.\)]", clean_search)
            or "ВІДПОВІДНО ДО" in clean_search.upper()
            or "ЗГІДНО З" in clean_search.upper()
            or "НАКАЗУЮ" in clean_search.upper()
            or "ПРИЗНАЧИТИ" in clean_search.upper()
            or "НАПРАВИТИ" in clean_search.upper()
            or "ВІДРЯДИТИ" in clean_search.upper()
            or "ЗВІЛЬНИТИ" in clean_search.upper()
            or "ВІЙСЬКОВОСЛУЖБОВЦІВ" in clean_search.upper()
        ):
            content_start_idx = idx
            break
        if clean:
            header_lines.append(line)

    # ── Компілюємо патерни для кожної ВЧ ──────────────────────────────────────
    unit_patterns: list[tuple[str, str, str, re.Pattern]] = []
    unit_abbr_map: dict[str, str] = {}

    canonical_key_map: dict[str, str] = {}
    corps_map: dict[str, str] = {}
    cipher_to_primary_key: dict[str, str] = {}
    entry_routes_by_id: dict[int, tuple[str, str]] = {}
    route_entries_by_sender_key: dict[str, dict] = {}

    # ── Прохід 1: визначаємо ЄДИНИЙ шифр для кожного Корпусу ─────────────────
    corps_resolved_cipher: dict[str, str] = {}  # corps_abbr -> шифр корпусу

    for open_name, mapped_val in mapping_dict.items():
        if not isinstance(mapped_val, dict):
            continue
        corps_col = str(mapped_val.get("corps", "")).strip()
        if corps_col:
            corps_abbr = _extract_corps_abbr(corps_col)
            corps_map[open_name] = corps_col
            if corps_abbr not in corps_resolved_cipher:
                corps_entry = _find_corps_entry(corps_col, corps_abbr, mapping_dict)
                if corps_entry:
                    corps_resolved_cipher[corps_abbr] = _short_closed_code(str(corps_entry["cipher"]))
                else:
                    # Якщо корпус не має окремого рядка — не підставляємо шифр підпорядкованої частини!
                    corps_resolved_cipher[corps_abbr] = ""

    # ── Прохід 2: компілюємо патерни з єдиним sender_key ─────────────────────
    # Сортуємо: пріоритетні рядки з # обробляються першими
    sorted_mapping = sorted(
        mapping_dict.items(),
        key=lambda item: (not str(item[0]).startswith("#"), -len(str(item[0]))),
    )

    for open_name, mapped_val in sorted_mapping:
        corps_col = ""
        if isinstance(mapped_val, dict):
            closed_code = str(
                mapped_val.get("cipher")
                or mapped_val.get("closed_name")
                or open_name
            )
            corps_col = str(mapped_val.get("corps", "")).strip()
            abbreviation = str(
                mapped_val.get("abbreviation")
                or mapped_val.get("abbr")
                or mapped_val.get("short_name")
                or mapped_val.get("скорочення")
                or mapped_val.get("скорочена_назва")
                or ""
            ).strip()
        else:
            closed_code = str(mapped_val)
            abbreviation = ""

        short_cipher = _short_closed_code(closed_code)
        sender_key, resolved_abbr = _build_sender_key(closed_code, abbreviation, corps_col, corps_resolved_cipher, mapping_dict)

        if short_cipher:
            if short_cipher in cipher_to_primary_key:
                primary_key = cipher_to_primary_key[short_cipher]
                canonical_key_map[sender_key] = primary_key
                sender_key = primary_key
            else:
                cipher_to_primary_key[short_cipher] = sender_key

        if isinstance(mapped_val, dict):
            entry_routes_by_id[id(mapped_val)] = (sender_key, open_name)
            route_entry = mapped_val
            if corps_col:
                corps_abbr = _extract_corps_abbr(corps_col)
                corps_entry = _find_corps_entry(corps_col, corps_abbr, mapping_dict)
                if isinstance(corps_entry, dict):
                    route_entry = corps_entry
            route_entries_by_sender_key[sender_key] = route_entry

        unit_abbr_map[sender_key] = resolved_abbr or unit_abbr_map.get(sender_key, "")

        for variant in [open_name, closed_code, f"в/ч {short_cipher}", short_cipher, abbreviation, corps_col]:
            if variant:
                canonical_key_map[variant] = sender_key
                if corps_col:
                    corps_abbr = _extract_corps_abbr(corps_col)
                    canonical_key_map[corps_abbr] = sender_key

        clean_open = open_name.lstrip("#").strip()
        low_open = clean_open.lower()
        is_tck_entry = "тцк" in low_open or "територіальн" in low_open or "центр" in low_open
        if fuzzy_match or is_tck_entry:
            pattern = _build_unit_fuzzy_pattern(clean_open)
        else:
            pattern = re.compile(re.escape(clean_open), re.IGNORECASE)

        unit_patterns.append((open_name, closed_code, corps_col, sender_key, pattern))

    unit_patterns.sort(key=lambda x: (not str(x[0]).startswith("#"), -len(str(x[0]))))

    # ── Районний/міський ТЦК ніколи не отримує власного витягу ───────────────
    # Правило: такий ТЦК ЗАВЖДИ прямує до свого обласного, а «Кому»/«Куди»
    # беруться з рядка області. Коли районного ТЦК у таблиці немає, це вже
    # працювало (`_extract_tck_sender` одразу давав область). Але якщо він там
    # є ОКРЕМИМ РЯДКОМ, він збігався як звичайна частина і на нього
    # створювався окремий витяг.
    def _is_tck_name(name: str) -> bool:
        low = str(name).casefold()
        return "тцк" in low or "територіальн" in low

    tck_redirect: dict[str, str] = {}
    for open_name, _code, _corps, sender_key, _pattern in unit_patterns:
        clean = str(open_name).lstrip("#").strip()
        low = clean.casefold()
        if not _is_tck_name(low) or "обласн" in low:
            continue
        oblast_name = _extract_tck_sender(clean)
        if not oblast_name:
            continue
        for other_name, _c, _cc, other_key, _p in unit_patterns:
            other_clean = str(other_name).lstrip("#").strip()
            if other_key == sender_key or "обласн" not in other_clean.casefold():
                continue
            if _extract_tck_sender(other_clean) == oblast_name:
                tck_redirect[sender_key] = other_key
                break

    if tck_redirect:
        unit_patterns = [
            (name, code, corps, tck_redirect.get(key, key), pattern)
            for (name, code, corps, key, pattern) in unit_patterns
        ]
        for old_key, new_key in tck_redirect.items():
            if new_key in route_entries_by_sender_key:
                route_entries_by_sender_key[old_key] = route_entries_by_sender_key[new_key]
        for alias, key in list(canonical_key_map.items()):
            if key in tck_redirect:
                canonical_key_map[alias] = tck_redirect[key]
        for entry_id, (key, name) in list(entry_routes_by_id.items()):
            if key in tck_redirect:
                entry_routes_by_id[entry_id] = (tck_redirect[key], name)

    def route_from_current_table(entry: dict | None) -> tuple[str, str] | None:
        """Повертає маршрут лише для рядка поточної Excel-таблиці."""
        if not isinstance(entry, dict):
            return None
        return entry_routes_by_id.get(id(entry))

    # Шукаємо вихідну ВЧ в усій преамбулі до першого справжнього пункту.
    # Межа `header/content` є технічною і в Word може пройти просто посеред
    # назви частини: «46 окремого» лишається у header, а
    # «ремонтно-відновлювального полку ЗВІЛЬНИТИ…» починає content. Пошук по
    # обох половинах окремо тоді гарантовано нічого не знаходить.
    first_item_idx = next(
        (
            idx
            for idx, line in enumerate(lines)
            if re.match(
                r"^\s*\d{1,3}(?:\.\d{1,3})*[\.\)](?:\s|$)",
                re.sub(r"\s+", " ", line),
            )
        ),
        len(lines),
    )
    header_text = "\n".join(lines[:first_item_idx])
    header_zvidky_unit: tuple[str, str] | None = None
    for open_name, closed_code, corps_col, sender_key, pattern in unit_patterns:
        m = pattern.search(header_text)
        if m:
            header_zvidky_unit = (sender_key, open_name)
            break

    # ── Парсимо тіло наказу у блоки (§-параграфи та пронумеровані пункти) ─────
    blocks = []
    current_parent_heading = ""
    current_block = None

    active_heading_start_idx = -1
    active_heading_end_idx = -1
    # Ієрархія шапки як список сегментів по рівнях: [§, шапка, підшапка, ...].
    # Кожен елемент — [start_line, end_line, text]. Генератор витягів
    # повторює лише той рівень, що реально змінився (§ і незмінна шапка не
    # дублюються для кожного пункту одного розділу — AGENT.md, розд. 4).
    heading_segments: list[list] = []
    # Стан ієрархії ДО останнього доданого/заміненого рівня (напр. підшапки
    # «У ЗАПАС ЗА ПІДПУНКТОМ …:») та прапорець, що цей останній рівень сам є
    # підставою звільнення. Потрібно для двох випадків:
    #  1) пункт із власною підставою звільнення в тексті не успадковує чужу
    #     зовнішню підшапку;
    #  2) сусідня підшапка-причина звільнення в межах ОДНОГО § заміняє
    #     попередню (той самий рівень), а не додає ще один рівень вкладеності.
    pre_discharge_heading_segments: list[list] | None = None
    last_segment_is_discharge = False
    previous_line_was_heading = False
    for rel_idx, line in enumerate(lines[content_start_idx:]):
        abs_idx = content_start_idx + rel_idx
        clean = line.strip()
        clean_search = re.sub(r"\s+", " ", clean)
        if not clean:
            previous_line_was_heading = False
            if current_block:
                current_block["lines"].append("")
                current_block["end_line"] = abs_idx
                if current_block["type"] == "section":
                    active_heading_end_idx = abs_idx
                    if heading_segments:
                        heading_segments[-1][1] = abs_idx
            continue

        # Підшапка-підстава звільнення не завжди закінчується двокрапкою:
        # «У ВІДСТАВКУ ЗА ПІДПУНКТОМ “б” (… про непридатність до військової
        # служби).» закінчується КРАПКОЮ. Такий рядок не впізнавався як
        # підшапка, поглинався сусіднім блоком і друкувався у витягу НЕ на
        # своєму місці — «підшапка переміщалася вниз». Пункт із власною
        # підставою сюди не потрапляє: він починається з номера.
        is_discharge_subheading = (
            _is_inline_discharge_basis(clean_search)
            and _starts_a_new_heading(clean_search)
        )
        is_section_marker = clean_search.startswith("§") or (
            (
                "відповідно до" in clean_search.casefold()
                or "згідно з" in clean_search.casefold()
                or clean_search.endswith(":")
                or "ВІЙСЬКОВОСЛУЖБОВЦІВ" in clean_search.upper()
                or is_discharge_subheading
            )
            and not re.match(r"^\d+[\.\d]*", clean_search)
            and not clean_search.casefold().startswith("підстава")
        )

        is_new_item = (
            not is_section_marker
            # Наявність «р.н.» наприкінці того самого абзацу не скасовує
            # номер пункту на його початку. Раніше через це пункти 2/4/7/8
            # приєднувалися до шапки й зовсім не потрапляли в маршрутизацію.
            and not clean.startswith("Підстава")
            and bool(re.match(r"^\d{1,3}(?:\.\d{1,3})*[\.\)]\s+", clean))
        )

        # Шапку розділу часто розбиває МʼЯКИЙ ПЕРЕНОС (Shift+Enter) просто
        # посеред назви частини: «…офіцерського складу 46 окремого⏎
        # ремонтно-відновлювального полку ЗВІЛЬНИТИ…». Обидва уламки
        # проходили як шапка (другий — бо закінчується двокрапкою) і ставали
        # ОКРЕМИМИ блоками, тож назва частини лишалася розірваною і не
        # знаходилась зовсім: усі пункти розділу втрачали адресата.
        #
        # Уламок продовження починається з МАЛОЇ літери — цим він і
        # відрізняється від справжньої підшапки («У ЗАПАС ЗА ПІДПУНКТОМ …:»),
        # яка завжди починається з великої.
        is_heading_continuation = (
            previous_line_was_heading
            and not clean.startswith("§")
            and not is_new_item
            and not clean.casefold().startswith("підстава")
            and not _starts_a_new_heading(clean)
            and current_block is not None
            and current_block["type"] == "section"
        )
        if is_heading_continuation:
            current_block["lines"].append(line)
            current_block["end_line"] = abs_idx
            active_heading_end_idx = abs_idx
            if heading_segments:
                heading_segments[-1][1] = abs_idx
                heading_segments[-1][2] = f"{heading_segments[-1][2]} {clean}".strip()
                current_parent_heading = "\n\n".join(seg[2] for seg in heading_segments)
                current_block["heading"] = current_parent_heading
                current_block["heading_ranges"] = [(seg[0], seg[1]) for seg in heading_segments]
            # Підтримуємо ланцюг із трьох і більше уламків. Середній уламок
            # не зобов'язаний сам закінчуватися двокрапкою або містити
            # «Відповідно до», але наступний рядок усе ще є продовженням
            # тієї самої шапки після Shift+Enter.
            previous_line_was_heading = True
            continue

        previous_line_was_heading = is_section_marker

        if is_section_marker:
            if clean.startswith("§") or not current_parent_heading:
                current_parent_heading = line
                active_heading_start_idx = abs_idx
                heading_segments = [[abs_idx, abs_idx, line]]
                pre_discharge_heading_segments = None
                last_segment_is_discharge = False
            else:
                is_discharge_line = _is_inline_discharge_basis(line)
                if is_discharge_line and last_segment_is_discharge and heading_segments:
                    # Сусідня підшапка-причина звільнення (напр. підпункт «в»
                    # після підпункту «а») заміняє попередню на тому самому
                    # рівні ієрархії, а не додається як глибша вкладеність.
                    pre_discharge_heading_segments = [list(seg) for seg in heading_segments[:-1]]
                    heading_segments[-1] = [abs_idx, abs_idx, line]
                else:
                    pre_discharge_heading_segments = [list(seg) for seg in heading_segments]
                    heading_segments.append([abs_idx, abs_idx, line])
                last_segment_is_discharge = is_discharge_line
                current_parent_heading = "\n\n".join(seg[2] for seg in heading_segments)
            active_heading_end_idx = abs_idx
            current_block = {
                "type": "section",
                "heading": current_parent_heading,
                "label": clean.split()[0] if clean.split() else "Розділ",
                "lines": [line],
                "start_line": abs_idx,
                "end_line": abs_idx,
                "heading_start_line": active_heading_start_idx,
                "heading_end_line": active_heading_end_idx,
                "heading_ranges": [(seg[0], seg[1]) for seg in heading_segments],
            }
            blocks.append(current_block)
        elif is_new_item:
            match = re.match(r"^(\d+[\.\d]*)", clean)
            label = f"Пункт {match.group(1)}" if match else "Пункт"
            item_heading = current_parent_heading
            item_segments = heading_segments
            if (
                last_segment_is_discharge
                and pre_discharge_heading_segments is not None
                and _is_inline_discharge_basis(clean)
            ):
                # Пункт сам формулює власну підставу звільнення — не
                # копіюємо зовнішню підшапку «У ЗАПАС ЗА ПІДПУНКТОМ …»,
                # лишаємо лише попередні рівні ієрархії (§ + основна шапка).
                item_segments = pre_discharge_heading_segments
                item_heading = "\n\n".join(seg[2] for seg in item_segments)
            item_heading_start = item_segments[0][0] if item_segments else active_heading_start_idx
            item_heading_end = (
                item_segments[-1][1] if item_segments
                else (active_heading_end_idx if active_heading_end_idx >= 0 else abs_idx - 1)
            )
            current_block = {
                "type": "item",
                "heading": item_heading,
                "label": label,
                # Ієрархія БЕЗ зовнішньої підшапки-підстави. Знадобиться, якщо
                # власна підстава пункту стоїть не в першому його рядку.
                "_pre_discharge_segments": (
                    [list(seg) for seg in pre_discharge_heading_segments]
                    if last_segment_is_discharge and pre_discharge_heading_segments is not None
                    else None
                ),
                "lines": [line],
                "start_line": abs_idx,
                "end_line": abs_idx,
                "heading_start_line": item_heading_start,
                "heading_end_line": item_heading_end,
                "heading_ranges": [(seg[0], seg[1]) for seg in item_segments],
            }
            blocks.append(current_block)
        else:
            if current_block:
                current_block["lines"].append(line)
                current_block["end_line"] = abs_idx
            else:
                current_block = {
                    "type": "item",
                    "heading": current_parent_heading,
                    "label": "Основний текст",
                    "lines": [line],
                    "start_line": abs_idx,
                    "end_line": abs_idx,
                    "heading_start_line": active_heading_start_idx,
                    "heading_end_line": active_heading_end_idx if active_heading_end_idx >= 0 else abs_idx - 1,
                }
                blocks.append(current_block)

    # ── Шукаємо ВЧ у кожному пункті ───────────────────────────────────────────
    unit_data_map: dict[str, dict] = {}
    unit_counts: dict[str, int] = {}
    unit_open_names: dict[str, set[str]] = {}
    match_report_rows = []
    # Ці два списки є джерелом даних для вкладок і Excel-контролю
    # генератора. Раніше ключі очікувалися в інтерфейсі, але не поверталися
    # звідси, через що «Контроль пропущених пунктів» завжди був порожнім.
    unmatched_items: list[dict] = []
    skipped_items: list[dict] = []
    routing_audit: list[dict] = []

    # Власна підстава звільнення може стояти НЕ в першому рядку пункту: текст
    # часто розриває мʼякий перенос, і «У ЗАПАС ЗА ПІДПУНКТОМ …» опиняється на
    # другому рядку. Перевірка при створенні блоку бачила лише перший рядок,
    # тому до пункту приліплювалася ЧУЖА зовнішня підшапка (правило 4 AGENT.md).
    for item_block in blocks:
        if item_block.get("type") != "item":
            continue
        alt_segments = item_block.pop("_pre_discharge_segments", None)
        if not alt_segments:
            continue
        alt_ranges = [(seg[0], seg[1]) for seg in alt_segments]
        if item_block.get("heading_ranges") == alt_ranges:
            continue
        if _is_inline_discharge_basis(chr(10).join(item_block.get("lines", []))):
            item_block["heading"] = (chr(10) * 2).join(seg[2] for seg in alt_segments)
            item_block["heading_ranges"] = alt_ranges

    processed_lines = list(lines)
    active_section_units: set[tuple[str, str]] = (
        {header_zvidky_unit} if header_zvidky_unit else set()
    )

    for block in blocks:
        if block["type"] == "section":
            section_raw_text = "\n".join(block["lines"])
            # `lines` містить лише фізичний уламок, який створив поточний
            # section-блок. Word може розбити одну шапку м'яким переносом,
            # окремими абзацами/комірками або навіть порожнім абзацом. Повна
            # актуальна ієрархія вже зібрана в `heading`; саме її треба
            # перевіряти під час пошуку адресата, інакше номер та початок
            # назви лишаються в одному блоці, а тип частини — в іншому.
            # Не вибираємо одне з двох джерел. `heading` містить успадковану
            # ієрархію шапок, а `lines` — усі фізичні абзаци поточного блоку.
            # У реальних DOCX продовження шапки може починатися з великої
            # літери: тоді воно залишається у `lines`, але не додається до
            # `heading`. Саме в такому продовженні часто стоїть назва ВЧ.
            heading_text = str(block.get("heading") or "").strip()
            section_search_text = "\n\n".join(
                part for part in (heading_text, section_raw_text.strip()) if part
            )
            sec_units: set[tuple[str, str]] = set()

            # Шукаємо ЗВІДКИ (вихідна ВЧ) та КУДИ (цільова ВЧ) у шапці розділу
            match_kudy = re.search(
                r"(?:призначити\s+до|направити\s+до|відрядити\s+до|у\s+розпорядження)\s+([^:\n\.]+)",
                section_raw_text,
                re.IGNORECASE,
            )
            kudy_text = match_kudy.group(1).strip() if match_kudy else ""

            # 1. Знаходимо ВЧ напрямку КУДИ
            kudy_units: set[tuple[str, str]] = set()
            if kudy_text and "цього саг" not in kudy_text.lower() and "цієї саг" not in kudy_text.lower() and "того ж" not in kudy_text.lower():
                for open_name, closed_code, corps_col, sender_key, pattern in unit_patterns:
                    m = pattern.search(kudy_text)
                    if m:
                        matched_str = m.group(0)
                        low_open = open_name.lower()
                        is_tck_entry = "тцк" in low_open or "територіальн" in low_open or "центр" in low_open
                        target_name = sender_key if is_tck_entry else (corps_col if (corps_col and ("корпус" in matched_str.lower() or "ак" in matched_str.lower())) else open_name)
                        kudy_units.add((sender_key, target_name))
                if not kudy_units:
                    sec_tck = _extract_tck_sender(kudy_text)
                    if sec_tck:
                        sec_tck_entry = _find_entry_in_mapping(sec_tck, sec_tck, mapping_dict)
                        sec_tck_route = route_from_current_table(sec_tck_entry)
                        if sec_tck_route:
                            kudy_units.add(sec_tck_route)

            # 2. Знаходимо ВЧ напрямку ЗВІДКИ (або загальну ВЧ у шапці).
            #
            # Шапка може називати КІЛЬКА частин: «По військовій частині А1111 та
            # військовій частині А2222:». Раніше цикл зупинявся на першій
            # знайденій (`break`), тож пункти під такою шапкою йшли лише в ОДИН
            # витяг — причому в який саме, залежало від порядку рядків у
            # словнику, тому виглядало випадковим. Напрямок КУДИ вище збирає всі
            # збіги; тут робимо так само.
            zvidky_units: set[tuple[str, str]] = set()
            for open_name, closed_code, corps_col, sender_key, pattern in unit_patterns:
                m = pattern.search(section_search_text)
                if m:
                    matched_str = m.group(0)
                    low_open = open_name.lower()
                    is_tck_entry = "тцк" in low_open or "територіальн" in low_open or "центр" in low_open
                    target_name = sender_key if is_tck_entry else (corps_col if (corps_col and ("корпус" in matched_str.lower() or "ак" in matched_str.lower())) else open_name)
                    zvidky_units.add((sender_key, target_name))

            sec_units = zvidky_units | kudy_units
            if sec_units:
                active_section_units = sec_units
            elif str(block["lines"][0] if block["lines"] else "").strip().startswith("§"):
                # Правило 4.3: fallback на вихідну ВЧ із шапки наказу діє для
                # НОВОГО розділу §, який сам частини не називає.
                active_section_units = {header_zvidky_unit} if header_zvidky_unit else set()
            # Інакше це ПІДШАПКА всередині того самого § («У ЗАПАС ЗА
            # ПІДПУНКТОМ …:»). Вона задає підставу, а не адресата, тому
            # частину, названу в шапці розділу, треба ЗБЕРЕГТИ. Раніше вона
            # скидалася, і всі пункти під підшапкою втрачали адресата.
            continue

        if block["type"] != "item":
            continue

        raw_text_clean = re.sub(
            r"\b(?:передов\w*|запасн\w*|головн\w*|)?\s*"
            r"(?:командн\w*\s+пункт\w*|пункт\w*\s+управління|ЗКП|ГКП|ППУ|КП)\s+",
            "",
            "\n".join(block["lines"]),
            flags=re.IGNORECASE | re.UNICODE,
        )
        full_item_text = "\n".join(block["lines"]).strip()
        block_raw_text = _get_item_main_text(block["lines"])
        block_replaced_lines = list(block["lines"])
        matched_units_in_block: set[tuple[str, str]] = set()
        item_destinations: set[tuple[str, str]] = set()
        matched_open_names_for_item: list[str] = []

        # 1. Зіставлення виключно за пошуковими назвами зі стовпця A.
        matched_clean_names = set()
        for open_name, closed_code, corps_col, sender_key, pattern in unit_patterns:
            clean_name = open_name.lstrip("#").strip().lower()
            if clean_name in matched_clean_names:
                continue

            all_matches = pattern.findall(block_raw_text)
            if not all_matches:
                continue

            matched_clean_names.add(clean_name)
            if open_name not in matched_open_names_for_item:
                matched_open_names_for_item.append(open_name)

            for found_form in set(str(m) if isinstance(m, str) else str(m[0]) for m in all_matches):
                existing = next(
                    (r for r in match_report_rows if r[0] == open_name and r[1] == found_form),
                    None,
                )
                if existing:
                    idx = match_report_rows.index(existing)
                    match_report_rows[idx] = (
                        existing[0], existing[1], existing[2], existing[3] + 1
                    )
                else:
                    match_report_rows.append((open_name, found_form, closed_code, 1))

            block_replaced_lines = [pattern.sub(closed_code, ln) for ln in block_replaced_lines]
            matched_str = str(all_matches[0]) if all_matches else ""
            low_open = open_name.lower()
            is_tck_entry = "тцк" in low_open or "територіальн" in low_open or "центр" in low_open
            target_name = sender_key if is_tck_entry else (corps_col if (corps_col and ("корпус" in matched_str.lower() or "ак" in matched_str.lower())) else open_name)
            item_destinations.add((sender_key, target_name))

        # 2. Центри рекрутингу: у словнику номер часто записано після назви
        # («Центр рекрутингу № 7»), а в наказі — перед нею («7 центру
        # рекрутингу»). Зіставляємо номер і ключові слова окремо, без
        # прив'язки до конкретного номера та лише серед записів Excel.
        recruiting_numbers = set(
            re.findall(
                r"\b(\d{1,3})[-\s]*(?:-?й|-?го|-?му|-?м)?\s*"
                r"(?:центр\w*|цнтр\w*)\s+рекрутинг\w*\b",
                block_raw_text,
                re.IGNORECASE | re.UNICODE,
            )
        )
        matched_recruiting_entries: set[str] = set()
        for open_name, closed_code, corps_col, sender_key, pattern in unit_patterns:
            entry_text = str(open_name).casefold()
            if open_name in matched_recruiting_entries:
                continue
            matching_number = next(
                (
                    number
                    for number in recruiting_numbers
                    if re.search(rf"(?<!\d){re.escape(number)}(?!\d)", entry_text)
                ),
                None,
            )
            is_recruiting_entry = (
                "центр" in entry_text and "рекрут" in entry_text
            ) or bool(
                matching_number
                and re.search(
                    rf"(?<!\d){re.escape(matching_number)}(?!\d)\s*цр\b",
                    entry_text,
                    re.IGNORECASE | re.UNICODE,
                )
            )
            if (
                not recruiting_numbers
                or not matching_number
                or not is_recruiting_entry
            ):
                continue
            matched_recruiting_entries.add(open_name)
            if open_name not in matched_open_names_for_item:
                matched_open_names_for_item.append(open_name)
            item_destinations.add((sender_key, sender_key))

        # 3. Якщо в тексті згадується ТЦК (районний/міський -> Область) — додаємо його
        for tck_sender in _extract_tck_senders(block_raw_text):
            if tck_sender in [s for s, _ in item_destinations]:
                continue
            tck_entry = _find_entry_in_mapping(tck_sender, tck_sender, mapping_dict)
            tck_route = route_from_current_table(tck_entry)
            if tck_route and tck_route[0] not in [s for s, _ in item_destinations]:
                item_destinations.add(tck_route)

        # 4. Якщо в таблиці немає відповідностей — перевіряємо чи в тексті явно є АК
        if not item_destinations:
            corps_name = _extract_army_corps(block_raw_text)
            if corps_name:
                corps_abbr = _extract_corps_abbr(corps_name)
                corps_entry = _find_corps_entry(corps_name, corps_abbr, mapping_dict)
                corps_route = route_from_current_table(corps_entry)
                if corps_route:
                    item_destinations.add(corps_route)

        # Якщо у тексті пункту є вказівка на внутрішнє переміщення ("цього самого батальйону", "цієї самої бригади", "цього ж полку"),
        # адресат береться з контексту. Але явно названа в цьому ж пункті ВЧ
        # або ТЦК є саме цим контекстом — її не можна відкидати. Через старе
        # очищення тут губилися, зокрема, пункти «... обласного ТЦК ... цього
        # самого центру».
        has_internal_ref = bool(
            re.search(
                r"\b(?:цього|цієї|того)\s+(?:самого|самої|ж)\b",
                block_raw_text,
                re.IGNORECASE,
            )
        )

        # Формуємо підсумковий набір отримувачів пункту (Джерело ЗВІДКИ + Призначення КУДИ)
        base_source = active_section_units or ({header_zvidky_unit} if header_zvidky_unit else set())
        matched_units_in_block = set(base_source) | item_destinations

        def audit_recipient_names(recipient_pairs: set[tuple[str, str]]) -> list[str]:
            return sorted(
                {
                    _normalize_key(raw_code, canonical_key_map) or str(open_name).strip()
                    for raw_code, open_name in recipient_pairs
                    if str(raw_code).strip() or str(open_name).strip()
                }
            )

        item_recipient_names = audit_recipient_names(item_destinations)
        context_recipient_names = audit_recipient_names(set(base_source))
        final_recipient_names = audit_recipient_names(matched_units_in_block)
        # Зміни на посади в управлінні не включаються до загального переліку
        # витягів. Це окремий напрямок розсилки, який користувач обробляє поза
        # цим генератором, тому такий пункт не є «пропущеним».
        is_management_change = (
            not final_recipient_names
            # Перевіряємо повний текст пункту, а не тільки розпорядчу його
            # частину: фільтр маршрутизації прибирає службові/облікові рядки,
            # але слово «управління» може бути саме в такому рядку.
            and bool(re.search(r"\bуправлінн\w*\b", full_item_text, re.IGNORECASE))
        )
        applied_rules = []
        if context_recipient_names:
            applied_rules.append("адресат із шапки розділу/наказу")
        if item_recipient_names:
            applied_rules.append("адресат знайдено в пункті")
        if has_internal_ref:
            applied_rules.append(
                "внутрішнє переміщення: "
                + ("підтверджено названим адресатом" if item_recipient_names else "адресат із контексту")
            )
        if is_management_change:
            applied_rules.append("зміна до управління: витяг виключено із загального переліку")
        routing_audit.append(
            {
                "label": block.get("label", ""),
                "matched_entries": ", ".join(matched_open_names_for_item) or "—",
                "applied_rules": "; ".join(applied_rules) or "адресата не знайдено",
                "item_recipients": ", ".join(item_recipient_names) or "—",
                "context_recipients": ", ".join(context_recipient_names) or "—",
                "final_recipients": ", ".join(final_recipient_names) or "—",
            }
        )

        if matched_units_in_block:
            # Правило 4.1: text — завжди з відкритими назвами (для витягів)
            # text_cipher — з шифрами (для повідомлень/рішень)
            full_item_text_cipher = "\n".join(block_replaced_lines).strip()
            # Дедуплікація: один пункт наказу додається лише ОДИН раз на кожен унікальний norm_code
            seen_norm_codes: set[str] = set()
            for raw_code, open_name in matched_units_in_block:
                norm_code = _normalize_key(raw_code, canonical_key_map)
                if norm_code in seen_norm_codes:
                    continue
                seen_norm_codes.add(norm_code)

                # Пошук адресата вже завершено за стовпцем A. Тут лише беремо
                # вихідні B–F саме з обраного рядка (для підпорядкованої
                # частини — з уже визначеного рядка корпусу).
                mapping_entry = route_entries_by_sender_key.get(norm_code)
                if mapping_entry is None:
                    mapping_entry = _find_entry_in_mapping(norm_code, open_name, mapping_dict)
                rec_to = ""
                dest_where = ""
                if mapping_entry:
                    rec_to = str(mapping_entry.get("recipient_to") or "").strip()
                    dest_where = str(mapping_entry.get("destination_where") or mapping_entry.get("dislocation") or "").strip()

                # Якщо поле 'Куди' не заповнено в Excel — встановлюємо маркер 'КУДИ' для ручної правки (виділяється червоним)
                if not dest_where:
                    dest_where = "КУДИ"

                unit_entry = unit_data_map.setdefault(
                    norm_code,
                    {
                        "unit_code": norm_code,
                        "open_name": open_name,
                        "abbreviation": unit_abbr_map.get(norm_code, ""),
                        "recipient_to": rec_to,
                        "destination_where": dest_where,
                        "header_lines": header_lines,
                        "items": [],
                    },
                )
                if not unit_entry.get("recipient_to") and rec_to:
                    unit_entry["recipient_to"] = rec_to
                if (not unit_entry.get("destination_where") or unit_entry.get("destination_where") == "КУДИ") and dest_where and dest_where != "КУДИ":
                    unit_entry["destination_where"] = dest_where
                if "abbreviation" not in unit_entry or not unit_entry["abbreviation"]:
                    unit_entry["abbreviation"] = unit_abbr_map.get(norm_code, "")
                unit_entry["items"].append(
                    {
                        "parent_heading": block["heading"],
                        "label": block["label"],
                        "text": full_item_text,
                        "text_cipher": full_item_text_cipher,
                        "source_start_line": block.get("start_line", 0),
                        "source_end_line": block.get("end_line", 0),
                        "heading_start_line": block.get("heading_start_line", 0),
                        "heading_end_line": block.get("heading_end_line", 0),
                        "heading_ranges": block.get("heading_ranges", []),
                    }
                )
                unit_counts[norm_code] = unit_counts.get(norm_code, 0) + 1
                unit_open_names.setdefault(norm_code, set()).add(open_name)
        elif is_management_change:
            skipped_items.append(
                {
                    "label": block.get("label", ""),
                    "text": "\n".join(block.get("lines", [])).strip(),
                    "reason": "Зміна до управління: витяг не створюється у загальному переліку.",
                }
            )
        else:
            unmatched_items.append(
                {
                    "label": block.get("label", ""),
                    "text": "\n".join(block.get("lines", [])).strip(),
                    "reason": "Не знайдено адресата ні в пункті, ні в контексті розділу/наказу.",
                }
            )

    # ── Будуємо вихідні таблиці ────────────────────────────────────────────────
    table_rows = []
    for closed_code, count in unit_counts.items():
        open_str = ", ".join(sorted(unit_open_names.get(closed_code, [])))
        entry = unit_data_map[closed_code]
        raw_labels = [item.get("label", "") for item in entry["items"]]
        range_labels = _format_item_numbers_range(raw_labels)
        abbr = unit_abbr_map.get(closed_code, "")
        short_code = _short_closed_code(closed_code, abbreviation=abbr)
        table_rows.append((short_code, open_str, range_labels, count))

    table = DataTable(
        ("Закрита назва (ВЧ)", "Відкрита назва", "Номери пунктів витягу", "Кількість пунктів"),
        tuple(table_rows),
        "Розпізнані військові частини",
    )
    match_report = DataTable(
        ("Шаблон пошуку", "Знайдена форма у тексті", "Замінено на (шифр)", "Кількість"),
        tuple(match_report_rows),
        "Звіт відмінкових збігів",
    )

    final_text = "\n".join(processed_lines)
    summary = f"Знайдено частин: {len(unit_counts)} · Загалом згадок: {sum(unit_counts.values())} · Відмінкових форм розпізнано: {len(match_report_rows)}"

    return {
        "processed_text": final_text,
        "units_table": table,
        "units_list": list(unit_counts.keys()),
        "unit_paragraphs": unit_data_map,
        "unit_abbr_map": unit_abbr_map,
        "match_report": match_report,
        "unmatched_items": unmatched_items,
        "skipped_items": skipped_items,
        "routing_audit": routing_audit,
        "preamble_recipient": header_zvidky_unit[1] if header_zvidky_unit else "",
        "summary": summary,
    }


@node(
    name="Аналіз відправників та пунктів наказу",
    category="Наказ",
    description=(
        "Аналізує текст відкритого наказу, знаходить усіх відправників / адресатів "
        "(військові частини, ТЦК та СП, штаби тощо) та прив'язує кожен пункт "
        "до відповідного відправника."
    ),
    type_id="builtin.order.analyze_senders",
    outputs={
        "sender_paragraphs": "dict",
        "senders_list": "list",
        "table": "DataTable",
        "summary": "str",
    },
)
def analyze_senders(
    text: str = "",
    mapping: dict | None = None,
    default_sender_prefix: str = "Відправник ",
) -> dict:
    """Аналізує відкритий наказ і формує відповідність [Відправник / ВЧ] -> [Список пунктів]."""
    res = map_military_units(text=text, mapping=mapping, default_prefix=default_sender_prefix)
    units_map = res.get("unit_paragraphs", {})
    abbr_map = res.get("unit_abbr_map", {})
    sender_paragraphs: dict[str, list[str]] = {}
    table_rows = []

    for sender, data in units_map.items():
        short_sender = sender
        if isinstance(data, dict) and "items" in data:
            items = [item["text"] for item in data["items"] if "text" in item]
            raw_labels = [item.get("label", "") for item in data["items"]]
            range_labels = _format_item_numbers_range(raw_labels)
        else:
            items = [str(data)]
            range_labels = "-"
        sender_paragraphs[sender] = items
        table_rows.append((short_sender, range_labels, len(items)))

    table = DataTable(
        ("Військова частина / Відправник", "Номери пунктів витягу", "Кількість пунктів"),
        tuple(table_rows),
        "Аналіз відправників та пунктів витягу",
    )
    summary = f"Виявлено відправників: {len(sender_paragraphs)} · Опрацьовано пунктів: {sum(len(v) for v in sender_paragraphs.values())}"
    return {
        "sender_paragraphs": sender_paragraphs,
        "senders_list": list(sender_paragraphs.keys()),
        "table": table,
        "summary": summary,
    }


@node(
    name="Розділити наказ за відправниками",
    category="Наказ",
    description=(
        "Формує окремі текстові витяги або блоки наказу для кожного розпізнаного відправника "
        "(ВЧ), зберігаючи шапку та преамбулу (§) оригінального наказу."
    ),
    type_id="builtin.order.split_by_senders",
    outputs={
        "blocks": "dict",
        "table": "DataTable",
        "senders_count": "int",
        "summary": "str",
    },
)
def split_by_senders(
    text: str = "",
    mapping: dict | None = None,
    header: str = "",
) -> dict:
    """Розділяє відкритий наказ на окремі блоки / витяги за відправниками."""
    res = map_military_units(text=text, mapping=mapping)
    units_map = res.get("unit_paragraphs", {})
    abbr_map = res.get("unit_abbr_map", {})
    blocks: dict[str, str] = {}
    table_rows = []

    for sender, data in units_map.items():
        header_text = header
        if not header_text and isinstance(data, dict):
            header_text = "\n".join(data.get("header_lines", []))

        short_sender = sender
        items_text = ""
        range_labels = "-"
        count = 0
        if isinstance(data, dict) and "items" in data:
            # Групуємо пункти під відповідними преамбулами/заголовками розділів (§ 1 або ЗВІЛЬНИТИ...)
            section_groups: dict[str, list[str]] = {}
            raw_labels = []
            for item in data.get("items", []):
                heading = item.get("parent_heading", "").strip()
                item_text = item.get("text", "").strip()
                label = item.get("label", "")
                if label:
                    raw_labels.append(label)
                if item_text:
                    section_groups.setdefault(heading, []).append(item_text)

            range_labels = _format_item_numbers_range(raw_labels)
            count = len(data.get("items", []))

            body_sections = []
            for heading, item_texts in section_groups.items():
                if heading:
                    body_sections.append(f"{heading}\n\n" + "\n\n".join(item_texts))
                else:
                    body_sections.append("\n\n".join(item_texts))

            items_text = "\n\n".join(body_sections)
        elif isinstance(data, str):
            items_text = data
            count = 1

        full_doc = f"{header_text}\n\n{items_text}".strip() if header_text else items_text.strip()
        blocks[sender] = full_doc
        table_rows.append((short_sender, range_labels, count))

    table = DataTable(
        ("Військова частина (Адресат)", "Номери пунктів витягу", "Кількість пунктів"),
        tuple(table_rows),
        "Перелік відправників та пунктів витягу",
    )
    summary = f"Створено витягів за відправниками: {len(blocks)}"
    return {
        "blocks": blocks,
        "table": table,
        "senders_count": len(blocks),
        "summary": summary,
    }


# ── Правила заміни займенниково-іменникових зворотів ───────────────────────────
# ── Звороти-посилання на раніше згадану частину ──────────────────────────────
#
# «цієї самої бригади», «цього самого батальйону», «у цій самій роті» тощо
# мають ставати «…військової частини» у ТОМУ Ж відмінку, але вже в жіночому
# роді, бо «військова частина» — жіночого роду. Відмінок визначається за
# формою означення (цей/той/зазначений/вказаний/даний), яка однозначна.
# Прийменник не захоплюється: давальний і місцевий дають однакову форму
# («військовій частині»), тому «у цьому самому батальйоні» коректно стає
# «у цій самій військовій частині» без окремого правила на прийменник.

_UNIT_PHRASE_CASE_FORMS = {
    "Н": "військова частина",
    "Р": "військової частини",
    "Д": "військовій частині",   # давальний і місцевий збігаються
    "З": "військову частину",
    "О": "військовою частиною",
}

# Форми означення в жіночому роді — те, що потрапляє у вихідний текст.
_DETERMINER_FEMININE = {
    "цей":        {"Н": "ця",        "Р": "цієї",       "Д": "цій",        "З": "цю",        "О": "цією"},
    "той":        {"Н": "та",        "Р": "тієї",       "Д": "тій",        "З": "ту",        "О": "тією"},
    "зазначений": {"Н": "зазначена", "Р": "зазначеної", "Д": "зазначеній", "З": "зазначену", "О": "зазначеною"},
    "вказаний":   {"Н": "вказана",   "Р": "вказаної",   "Д": "вказаній",   "З": "вказану",   "О": "вказаною"},
    "даний":      {"Н": "дана",      "Р": "даної",      "Д": "даній",      "З": "дану",      "О": "даною"},
}

_SAM_FEMININE = {"Н": "сама", "Р": "самої", "Д": "самій", "З": "саму", "О": "самою"}

# Усі форми означення, які розпізнаються у вхідному тексті (будь-який рід).
_DETERMINER_SURFACE_FORMS = {
    "цей": {
        "Н": ("цей", "ця", "це"),
        "Р": ("цього", "цієї"),
        "Д": ("цьому", "цій"),
        "З": ("цю",),
        "О": ("цим", "цією"),
    },
    "той": {
        "Н": ("той", "та", "те"),
        "Р": ("того", "тієї", "тої"),
        "Д": ("тому", "тій"),
        "З": ("ту",),
        "О": ("тим", "тією", "тою"),
    },
    "зазначений": {
        "Н": ("зазначений", "зазначена", "зазначене"),
        "Р": ("зазначеного", "зазначеної"),
        "Д": ("зазначеному", "зазначеній"),
        "З": ("зазначену",),
        "О": ("зазначеним", "зазначеною"),
    },
    "вказаний": {
        "Н": ("вказаний", "вказана", "вказане"),
        "Р": ("вказаного", "вказаної"),
        "Д": ("вказаному", "вказаній"),
        "З": ("вказану",),
        "О": ("вказаним", "вказаною"),
    },
    "даний": {
        "Н": ("даний", "дана", "дане"),
        "Р": ("даного", "даної"),
        "Д": ("даному", "даній"),
        "З": ("дану",),
        "О": ("даним", "даною"),
    },
}

_DETERMINER_LOOKUP = {
    form: (lemma, case_label)
    for lemma, by_case in _DETERMINER_SURFACE_FORMS.items()
    for case_label, forms in by_case.items()
    for form in forms
}

# Ці форми поза контекстом надто неоднозначні («та» — сполучник «і»,
# «це»/«те» — вказівні частки), тому перетворюються лише разом зі словом «сам…».
_DETERMINER_REQUIRES_SAM = frozenset({"та", "те", "це"})

_SAM_SURFACE_FORMS = (
    "самий", "сама", "саме", "самого", "самої", "самому",
    "самій", "саму", "самим", "самою",
)

# Родові назви, на які може посилатися зворот. Сюди входять ЛИШЕ окремі
# військові частини та великі установи, які мають власний шифр (батальйон,
# бригада, центр, база, госпіталь тощо). Підрозділи всередині частини —
# рота, взвод, батарея, група, ескадрилья, відділ, відділення, служба,
# ланка, екіпаж, штаб, командний пункт — свідомо НЕ включені: вони не є
# військовою частиною, тому «цієї самої роти» лишається без змін.
# Назва потрібна як запобіжник: без неї зворот на кшталт «цього самого
# офіцера» помилково перетворювався б на військову частину.
_MASC_UNIT_STEMS = (
    "батальйон", "полк", "дивізіон", "загін", "загон", "корпус", "центр",
    "вузол", "вузл", "полігон", "госпіталь", "госпітал", "арсенал",
    "інститут", "університет",
)
_MASC_UNIT_ENDINGS = ("", "а", "у", "ю", "ом", "ем", "і", "ові", "еві", "я")

_FEM_UNIT_STEMS = (
    "бригад", "баз", "частин", "дивізі", "армі", "комендатур",
    "академі", "флотилі",
)
_FEM_UNIT_ENDINGS = ("а", "и", "і", "у", "ою", "я", "ї", "ю", "єю", "ею")

_NEUT_UNIT_STEMS = (
    "управлінн", "з'єднанн", "командуванн", "училищ", "угрупованн",
)
_NEUT_UNIT_ENDINGS = ("я", "ю", "ям", "і", "е", "ем", "а")

_UNIT_HEAD_NOUN_FORMS = sorted(
    {stem + ending for stem in _MASC_UNIT_STEMS for ending in _MASC_UNIT_ENDINGS}
    | {stem + ending for stem in _FEM_UNIT_STEMS for ending in _FEM_UNIT_ENDINGS}
    | {stem + ending for stem in _NEUT_UNIT_STEMS for ending in _NEUT_UNIT_ENDINGS},
    key=len,
    reverse=True,
)

_UNIT_REFERENCE_PATTERN = re.compile(
    r"\b(" + "|".join(re.escape(f) for f in sorted(_DETERMINER_LOOKUP, key=len, reverse=True)) + r")"
    r"(?:\s+(ж|же))?"
    r"(?:\s+(" + "|".join(re.escape(f) for f in sorted(_SAM_SURFACE_FORMS, key=len, reverse=True)) + r"))?"
    r"\s+(?:" + "|".join(re.escape(f) for f in _UNIT_HEAD_NOUN_FORMS) + r")\b"
    # ТЦК не є військовою частиною і в повідомленнях лишається відкритою
    # назвою, тому «центру комплектування…» не перетворюємо.
    r"(?!\s+комплектуванн)",
    re.IGNORECASE | re.UNICODE,
)


def _replace_unit_reference_phrase(match: re.Match) -> str:
    """Перетворює зворот-посилання на «…військової частини» у тому ж відмінку."""
    determiner_raw = match.group(1).lower()
    particle = match.group(2)
    sam_raw = match.group(3)

    lookup = _DETERMINER_LOOKUP.get(determiner_raw)
    if not lookup:
        return match.group(0)
    lemma, case_label = lookup

    # Неоднозначні короткі форми чіпаємо лише за наявності слова «сам…».
    if determiner_raw in _DETERMINER_REQUIRES_SAM and not sam_raw:
        return match.group(0)

    parts = [_DETERMINER_FEMININE[lemma][case_label]]
    if particle:
        parts.append(particle.lower())
    if sam_raw:
        parts.append(_SAM_FEMININE[case_label])
    parts.append(_UNIT_PHRASE_CASE_FORMS[case_label])

    result = " ".join(parts)
    # Зберігаємо велику літеру, якщо зворот стояв на початку речення
    # (_match_case у викликачів обробляє лише суцільний ВЕРХНІЙ РЕГІСТР).
    if match.group(0)[:1].isupper():
        result = result[:1].upper() + result[1:]
    return result


_UNIT_PHRASE_REPLACEMENTS = [
    (_UNIT_REFERENCE_PATTERN, _replace_unit_reference_phrase),
]


def is_tck_entry(mapped_val: dict | str) -> bool:
    """Чи є рядок словника територіальним центром комплектування (ТЦК).

    ТЦК не є закритою частиною: у змісті його назва лишається ПОВНОЮ
    відкритою (розд. 9.5.6), тому шифрування його рядки просто оминає.
    """
    if isinstance(mapped_val, dict):
        parts = (
            mapped_val.get("open_name"),
            mapped_val.get("cipher"),
            mapped_val.get("closed_name"),
            mapped_val.get("abbreviation"),
        )
    else:
        parts = (mapped_val,)
    haystack = " ".join(str(part or "") for part in parts).lower()
    return "тцк" in haystack or "комплектування" in haystack


def _format_full_closed_unit_text(mapped_val: dict | str, mapping_dict: dict) -> str:
    """Форматує закриту назву ВЧ: 'військової частини АXXXX', а якщо вказано корпус — 'військової частини АXXXX військової частини АYYYY'. Для ТЦК зберігає назву без змін."""
    if isinstance(mapped_val, dict):
        open_name = str(mapped_val.get("open_name", "")).strip()
        cipher = str(mapped_val.get("cipher") or mapped_val.get("closed_name") or "").strip()
        corps_name = str(mapped_val.get("corps", "")).strip()
    else:
        open_name = ""
        cipher = str(mapped_val).strip()
        corps_name = ""

    # Перевірка на ТЦК (територіальний центр комплектування) — залишаємо як є
    if is_tck_entry(mapped_val):
        return cipher or open_name

    def _to_unit_phrase(raw: str) -> str:
        clean = raw.strip()
        if not clean:
            return ""
        if clean.lower().startswith("військової частини") or clean.lower().startswith("військова частина"):
            return clean
        if clean.lower().startswith("в/ч"):
            code = clean[3:].strip()
            return f"військової частини {code}"
        if clean.upper().startswith("А") or clean.upper().startswith("A") or clean.isdigit():
            return f"військової частини {clean}"
        return f"військової частини {clean}"

    unit_code = _to_unit_phrase(cipher)

    if corps_name:
        # Стовпець D — це ПОСИЛАННЯ на рядок корпусу, і в ньому майже завжди
        # стоїть СКОРОЧЕННЯ («22 АК»), а не стовпець A того рядка
        # («22 армійський корпус»). Прямий `mapping_dict.get(corps_name)`
        # такого посилання не розвʼязував, і в текст підставлялося саме
        # скорочення: «військової частини А9999 військової частини 22 АК».
        # Далі це скорочення знаходив ВЛАСНИЙ патерн корпусу й шифрував ще
        # раз, через що виникали ланцюги «…22 АК військової частини А2222»
        # та дублі «військової частини військової частини».
        # Посилання розвʼязуємо тим самим `_find_corps_entry`, що й витяги.
        corps_entry = _find_corps_entry(corps_name, _extract_corps_abbr(corps_name), mapping_dict)
        if corps_entry:
            corps_cipher = str(corps_entry.get("cipher") if isinstance(corps_entry, dict) else corps_entry)
            corps_code = _to_unit_phrase(corps_cipher)
        else:
            corps_code = _to_unit_phrase(corps_name)
        return f"{unit_code} {corps_code}"

    return unit_code


def _match_case(original_matched_text: str, replacement_text: str) -> str:
    """Якщо оригінальний співпавший текст введений великими літерами (ALL CAPS), повертає заміну у ВЕЛИКИХ ЛІТЕРАХ."""
    if not original_matched_text or not replacement_text:
        return replacement_text
    letters = [ch for ch in original_matched_text if ch.isalpha()]
    if letters and all(ch.isupper() for ch in letters):
        return replacement_text.upper()
    return replacement_text


def _apply_custom_rules(text: str, rules: dict | list | str | None) -> tuple[str, int]:
    """Застосовує додаткові правила/виправлення, передані через нижній порт 'rules' / 'corrections'."""
    if not text or not rules:
        return text, 0

    count = 0
    rule_dict = {}

    if isinstance(rules, dict):
        rule_dict = dict(rules)
    elif isinstance(rules, list):
        for item in rules:
            if isinstance(item, dict):
                rule_dict.update(item)
            elif isinstance(item, str) and "->" in item:
                parts = item.split("->", 1)
                rule_dict[parts[0].strip()] = parts[1].strip()
    elif isinstance(rules, str):
        for line in rules.splitlines():
            if "->" in line:
                parts = line.split("->", 1)
                rule_dict[parts[0].strip()] = parts[1].strip()

    for old_val, new_val in rule_dict.items():
        if not old_val or not isinstance(new_val, str):
            continue
        pattern = re.compile(re.escape(str(old_val)), re.IGNORECASE)
        if pattern.search(text):
            matches = pattern.findall(text)
            count += len(matches)
            text = pattern.sub(lambda m: _match_case(m.group(0), new_val), text)

    return text, count


@node(
    name="Правила та виправлення (підключення знизу)",
    category="Наказ",
    description=(
        "Створює пакет додаткових правил та виправлень для підключення до нижнього порту 'rules' / 'corrections' "
        "нод опрацювання наказів. Підтримує виправлення у форматі 'старе -> нове'."
    ),
    type_id="builtin.order.create_rules",
    outputs={
        "rules": "dict",
        "count": "int",
        "summary": "str",
    },
)
def create_order_rules(
    text_rules: str = "",
    overrides: dict | None = None,
) -> dict:
    rule_dict = {}
    if overrides:
        rule_dict.update(overrides)
    if text_rules.strip():
        for line in text_rules.splitlines():
            if "->" in line:
                parts = line.split("->", 1)
                rule_dict[parts[0].strip()] = parts[1].strip()
    summary = f"Сформовано правил/виправлень: {len(rule_dict)}"
    return {
        "rules": rule_dict,
        "count": len(rule_dict),
        "summary": summary,
    }


@node(
    name="Генерація наказу про прийняття рішень (закритий)",
    category="Наказ",
    description=(
        "Генерує закритий наказ про прийняття рішень: видаляє/замінює відкриту шапку на закриту, "
        "конвертує всі відкриті назви частин у форматовані шифри ('військової частини АXXXX' чи 'військової частини АXXXX військової частини АYYYY' для корпусів), "
        "зберігає CAPS якщо оригінальний текст введений ВЕЛИКИМИ ЛІТЕРАМИ, замінює звороти ('цієї самої бригади' -> 'цієї самої військової частини'), "
        "застосовує виправлення з нижнього порту 'rules', та повертає таблицю виявлених частин і корпусів."
    ),
    type_id="builtin.order.generate_decision_order",
    execution_inputs=("exec",),
    execution_outputs=("then",),
    outputs={
        "decision_text": "str",
        "table": "DataTable",
        "replaced_count": "int",
        "summary": "str",
    },
)
def generate_decision_order(
    text: str = "",
    mapping: dict | None = None,
    new_header: str = "НАКАЗ командира військової частини А0000 (по стройовій частині)",
    fuzzy_match: bool = True,
    rules: dict | list | str | None = None,
) -> dict:
    """Генерує закритий наказ про прийняття рішень."""
    if not text.strip():
        return {
            "decision_text": "",
            "table": DataTable(("Відкрита назва", "Закритий код ВЧ", "Армійський корпус", "Форматована заміна"), ()),
            "replaced_count": 0,
            "summary": "Порожній текст наказу",
        }

    text = _fix_military_typos(text)
    mapping_raw = mapping or {}
    mapping_dict = {}
    for k, v in mapping_raw.items():
        clean_k = _fix_military_typos(str(k))
        if isinstance(v, dict):
            clean_v = dict(v)
            if "open_name" in clean_v:
                clean_v["open_name"] = _fix_military_typos(str(clean_v["open_name"]))
            mapping_dict[clean_k] = clean_v
        else:
            mapping_dict[clean_k] = v

    lines = [line.rstrip() for line in text.splitlines()]

    # 1. Знаходимо початок змістовної частини (після шапки наказу)
    content_start_idx = 0
    for idx, line in enumerate(lines):
        clean = line.strip()
        if (
            clean.startswith("§")
            or re.match(r"^\d+[\.\)]", clean)
            or "НАКАЗУЮ" in clean.upper()
            or "ПРИЗНАЧИТИ" in clean.upper()
            or "НАПРАВИТИ" in clean.upper()
            or "ВІДРЯДИТИ" in clean.upper()
            or "ЗВІЛЬНИТИ" in clean.upper()
            or "ВІЙСЬКОВОСЛУЖБОВЦІВ" in clean.upper()
        ):
            content_start_idx = idx
            break

    body_lines = lines[content_start_idx:]
    body_text = "\n".join(body_lines)

    replaced_count = 0
    report_rows = []

    # 2. Замінюємо відкриті назви частин на закриті формовані назви (із корпусом) та збереженням CAPS
    for open_name, mapped_val in mapping_dict.items():
        closed_code = _format_full_closed_unit_text(mapped_val, mapping_dict)
        if isinstance(mapped_val, dict):
            raw_cipher = str(mapped_val.get("cipher", ""))
            corps_info = str(mapped_val.get("corps", ""))
            abbreviation = str(mapped_val.get("abbreviation", "")).strip()
        else:
            raw_cipher = str(mapped_val)
            corps_info = ""
            abbreviation = ""

        matched = False

        # Зіставляємо за сигнатурою відкриту назву
        pattern = _build_unit_fuzzy_pattern(open_name) if fuzzy_match else re.compile(re.escape(open_name), re.IGNORECASE)
        matches = pattern.findall(body_text)
        if matches:
            replaced_count += len(matches)
            body_text = pattern.sub(lambda m: _match_case(m.group(0), closed_code), body_text)
            matched = True

        # Зіставляємо скорочення
        if abbreviation and abbreviation != open_name:
            abbr_pattern = _build_unit_fuzzy_pattern(abbreviation) if fuzzy_match else re.compile(re.escape(abbreviation), re.IGNORECASE)
            abbr_matches = abbr_pattern.findall(body_text)
            if abbr_matches:
                replaced_count += len(abbr_matches)
                body_text = abbr_pattern.sub(lambda m: _match_case(m.group(0), closed_code), body_text)
                matched = True

        if matched:
            report_rows.append((open_name, raw_cipher, corps_info or "—", closed_code))

    # 3. Замінюємо звороти "цієї самої бригади", "цього самого полку" тощо (із збереженням CAPS)
    for pattern, replacer in _UNIT_PHRASE_REPLACEMENTS:
        matches = pattern.findall(body_text)
        if matches:
            replaced_count += len(matches)
            def _make_phrase_rep(r_func):
                return lambda m: _match_case(m.group(0), r_func(m))
            body_text = pattern.sub(_make_phrase_rep(replacer), body_text)

    # 3.1. Нормалізація повторів "військової частини" та однакових шифрів
    # Згортання подвійних/потрійних "військової частини військової частини"
    body_text = re.sub(
        r"\b(?:військов(?:ої|а|у|ій|ою)\s+частин(?:и|а|у|і|ою)\s*){2,}",
        "військової частини ",
        body_text,
        flags=re.IGNORECASE,
    )
    body_text = re.sub(
        r"\b(?:ВІЙСЬКОВ(?:ОЇ|А|У|ІЙ|ОЮ)\s+ЧАСТИН(?:И|А|У|І|ОЮ)\s*){2,}",
        "ВІЙСЬКОВОЇ ЧАСТИНИ ",
        body_text,
    )
    # Згортання однакових повторів шифрів підряд: "військової частини А1111 військової частини А1111"
    body_text = re.sub(
        r"\b(військов\w+\s+частин\w+\s+[АA]\d+)(?:\s+\1\b)+",
        r"\1",
        body_text,
        flags=re.IGNORECASE,
    )
    body_text = re.sub(
        r"\b(ВІЙСЬКОВ\w+\s+ЧАСТИН\w+\s+[АA]\d+)(?:\s+\1\b)+",
        r"\1",
        body_text,
    )

    # 4. Застосовуємо додаткові правила/виправлення з нижнього порту 'rules'
    if rules:
        body_text, custom_count = _apply_custom_rules(body_text, rules)
        replaced_count += custom_count

    # 4.1. Повторна санітизація після користувацьких правил
    body_text = re.sub(
        r"\b(?:військов(?:ої|а|у|ій|ою)\s+частин(?:и|а|у|і|ою)\s*){2,}",
        "військової частини ",
        body_text,
        flags=re.IGNORECASE,
    )
    body_text = re.sub(
        r"\b(?:ВІЙСЬКОВ(?:ОЇ|А|У|ІЙ|ОЮ)\s+ЧАСТИН(?:И|А|У|І|ОЮ)\s*){2,}",
        "ВІЙСЬКОВОЇ ЧАСТИНИ ",
        body_text,
    )

    # 4.2. Гарантуємо рівно 1 порожній рядок (ентер) перед кожним пунктом наказу та рівно 2 перед підписантом
    body_lines_cleaned = []
    for idx, b_line in enumerate(body_text.splitlines()):
        b_clean = b_line.strip()
        is_item = bool(re.match(r"^\d{1,3}(?:\.\d{1,3})*[\.\)]\s+", b_clean))
        is_signer = bool(_ORDER_SIGNER_START_RE.match(b_clean))

        if is_signer and idx > 0 and body_lines_cleaned:
            while body_lines_cleaned and body_lines_cleaned[-1].strip() == "":
                body_lines_cleaned.pop()
            body_lines_cleaned.append("")
            body_lines_cleaned.append("")
        elif is_item and idx > 0 and body_lines_cleaned:
            while len(body_lines_cleaned) > 1 and body_lines_cleaned[-1].strip() == "" and body_lines_cleaned[-2].strip() == "":
                body_lines_cleaned.pop()
            if body_lines_cleaned[-1].strip() != "":
                body_lines_cleaned.append("")
        body_lines_cleaned.append(b_line)
    body_text = "\n".join(body_lines_cleaned)

    # 5. Збираємо фінальний закритий наказ (нова шапка + закрите тіло)
    header_str = new_header.strip() if new_header.strip() else ""
    final_text = f"{header_str}\n\n{body_text}".strip() if header_str else body_text.strip()
    table = DataTable(
        ("Відкрита назва", "Закритий код ВЧ", "Армійський корпус", "Форматована заміна у тексті"),
        tuple(report_rows),
        "Звіт частин та корпусів наказу",
    )
    summary = f"Сформовано закритий наказ. Знайдено ВЧ/корпусів: {len(report_rows)}, замін: {replaced_count}"

    return {
        "decision_text": final_text,
        "table": table,
        "replaced_count": replaced_count,
        "summary": summary,
    }


# ── БЛОЧНИЙ КОНСТРУКТОР НАКАЗУ ───────────────────────────────────────────────

@node(
    name="Розібрати наказ на блоки (Конструктор)",
    category="Наказ",
    description=(
        "Розбирає текст наказу на структуровані блоки-конструктори: шапка (header), "
        "розділи (§ / section), пронумеровані пункти (item), підстави (basis) та підписи (footer)."
    ),
    type_id="builtin.order.parse_to_blocks",
    execution_inputs=("exec",),
    execution_outputs=("then",),
    outputs={
        "blocks": "List",
        "table": "DataTable",
        "count": "int",
        "summary": "str",
    },
)
def parse_to_blocks(text: str = "") -> dict:
    if not text.strip():
        return {
            "blocks": [],
            "table": DataTable(("ID", "Тип", "Мітка", "Вміст"), ()),
            "count": 0,
            "summary": "Порожній текст наказу",
        }

    lines = [line.rstrip() for line in text.splitlines()]
    blocks = []

    order_action_keywords = (
        "НАКАЗУЮ", "ПРИЗНАЧИТИ", "ЗВІЛЬНИТИ", "УКЛАСТИ", "ПРОДОВЖИТИ",
        "ПРИСВОЇТИ", "ОГОЛОСИТИ", "ПРИЙНЯТИ", "ЗАРАХУВАТИ", "ВВАЖАТИ",
        "НАПРАВИТИ", "ВІДРЯДИТИ", "ВИКЛЮЧИТИ", "ІМЕНУВАТИ", "СКАСУВАТИ", "ПОНОВИТИ"
    )

    # 1. Знаходимо шапку наказу (до першого § або пункту або дієслова-команди)
    content_start_idx = len(lines)
    header_lines = []
    for idx, line in enumerate(lines):
        clean = line.strip()
        if (
            clean.startswith("§")
            or re.match(r"^\d+[\.\)]", clean)
            or any(kw in clean.upper() for kw in order_action_keywords)
        ):
            content_start_idx = idx
            break
        if clean:
            header_lines.append(line)

    if header_lines:
        blocks.append({
            "id": "block_0",
            "type": "header",
            "label": "Шапка наказу",
            "text": "\n".join(header_lines),
            "lines": header_lines,
        })

    # 2. Розбираємо тіло наказу
    current_parent_heading = ""
    current_block = None
    block_counter = len(blocks)

    for line in lines[content_start_idx:]:
        clean = line.strip()
        if not clean:
            if current_block:
                current_block["lines"].append("")
            continue

        is_section = clean.startswith("§") or (
            ("Відповідно до" in clean or "Згідно з" in clean or clean.endswith(":"))
            and not re.match(r"^\d+[\.\d]*", clean)
            and not clean.startswith("Підстава")
        )
        is_basis = clean.startswith("Підстава") or clean.startswith("Підстави")
        is_item = (
            not is_section
            and not is_basis
            and not ("р.н." in clean or "р. н." in clean)
            and bool(re.match(r"^\d{1,3}(?:\.\d{1,3})*[\.\)]\s+", clean))
        )
        is_footer = clean.startswith("Командир") or clean.startswith("ТВО") or clean.startswith("ТИМЧАСОВО") or clean.startswith("Згідно з оригіналом")

        if is_section:
            if clean.startswith("§") or not current_parent_heading:
                current_parent_heading = line
            else:
                current_parent_heading += "\n\n" + line
            current_block = {
                "id": f"block_{block_counter}",
                "type": "section",
                "label": clean.split()[0] if clean.split() else "Розділ",
                "heading": current_parent_heading,
                "text": line,
                "lines": [line],
            }
            block_counter += 1
            blocks.append(current_block)
        elif is_item:
            match = re.match(r"^(\d+[\.\d]*)", clean)
            label = f"Пункт {match.group(1)}" if match else "Пункт"
            current_block = {
                "id": f"block_{block_counter}",
                "type": "item",
                "label": label,
                "heading": current_parent_heading,
                "text": line,
                "lines": [line],
            }
            block_counter += 1
            blocks.append(current_block)
        elif is_basis:
            current_block = {
                "id": f"block_{block_counter}",
                "type": "basis",
                "label": "Підстава",
                "heading": current_parent_heading,
                "text": line,
                "lines": [line],
            }
            block_counter += 1
            blocks.append(current_block)
        elif is_footer:
            current_block = {
                "id": f"block_{block_counter}",
                "type": "footer",
                "label": "Підпис / Фінал",
                "heading": "",
                "text": line,
                "lines": [line],
            }
            block_counter += 1
            blocks.append(current_block)
        else:
            if current_block:
                current_block["lines"].append(line)
                current_block["text"] = "\n".join(current_block["lines"])
            else:
                current_block = {
                    "id": f"block_{block_counter}",
                    "type": "item",
                    "label": "Основний текст",
                    "heading": current_parent_heading,
                    "text": line,
                    "lines": [line],
                }
                block_counter += 1
                blocks.append(current_block)

    table_rows = [
        (b["id"], b["type"], b["label"], b["text"][:80] + ("..." if len(b["text"]) > 80 else ""))
        for b in blocks
    ]
    table = DataTable(("ID", "Тип", "Мітка", "Вміст (прев'ю)"), tuple(table_rows), "Блоки наказу")

    return {
        "blocks": blocks,
        "table": table,
        "count": len(blocks),
        "summary": f"Розділено наказ на {len(blocks)} блоків-конструкторів",
    }


@node(
    name="Трансформація та фільтрація блоків",
    category="Наказ",
    description=(
        "Фільтрує або змінює окремі блоки наказу: застосовує карту замін (відкриті -> 'військової частини АXXXX' / 'військової частини АXXXX військової частини АYYYY'), "
        "замінює звороти 'цієї самої бригади' на 'цієї самої військової частини', та фільтрує за типом блоків."
    ),
    type_id="builtin.order.filter_transform_blocks",
    execution_inputs=("exec",),
    execution_outputs=("then",),
    outputs={
        "blocks": "List",
        "table": "DataTable",
        "modified_count": "int",
        "summary": "str",
    },
)
def filter_transform_blocks(
    blocks: list | None = None,
    include_types: str = "",
    mapping: dict | None = None,
    replace_unit_phrases: bool = True,
) -> dict:
    if not blocks:
        return {"blocks": [], "table": DataTable(("ID", "Тип", "Мітка", "Вміст"), ()), "modified_count": 0, "summary": "Порожній список блоків"}

    mapping_dict = mapping or {}
    types_set = {t.strip().casefold() for t in include_types.split(",") if t.strip()} if include_types.strip() else None

    result_blocks = []
    modified_count = 0

    for block in blocks:
        if not isinstance(block, dict):
            continue
        b_type = str(block.get("type", "")).casefold()
        if types_set and b_type not in types_set:
            continue

        b_copy = dict(block)
        lines = list(b_copy.get("lines", []))
        text = b_copy.get("text", "\n".join(lines))
        orig_text = text

        # 1. Заміна назв ВЧ та корпусів на форматовані шифри
        for open_name, mapped_val in mapping_dict.items():
            closed_code = _format_full_closed_unit_text(mapped_val, mapping_dict)
            pattern = _build_unit_fuzzy_pattern(open_name)
            if pattern.search(text):
                text = pattern.sub(closed_code, text)

        # 2. Заміна зворотів ("цієї самої бригади" -> "цієї самої військової частини")
        if replace_unit_phrases:
            for pattern, replacer in _UNIT_PHRASE_REPLACEMENTS:
                if pattern.search(text):
                    text = pattern.sub(replacer, text)

        if text != orig_text:
            modified_count += 1
            b_copy["text"] = text
            b_copy["lines"] = text.splitlines()

        result_blocks.append(b_copy)

    table_rows = [
        (b.get("id", ""), b.get("type", ""), b.get("label", ""), b.get("text", "")[:80] + ("..." if len(b.get("text", "")) > 80 else ""))
        for b in result_blocks
    ]
    table = DataTable(("ID", "Тип", "Мітка", "Вміст (прев'ю)"), tuple(table_rows), "Трансформовані блоки")

    summary = f"Опрацьовано блоків: {len(result_blocks)} (змінено: {modified_count})"
    return {"blocks": result_blocks, "table": table, "modified_count": modified_count, "summary": summary}


@node(
    name="Зібрати наказ з блоків",
    category="Наказ",
    description=(
        "Збирає структурований список блоків-конструкторів у єдиний готовий текст наказу "
        "із можливістю заміни шапки чи розділювачів."
    ),
    type_id="builtin.order.assemble_from_blocks",
    execution_inputs=("exec",),
    execution_outputs=("then",),
    outputs={
        "text": "str",
        "count": "int",
        "summary": "str",
    },
)
def assemble_from_blocks(
    blocks: list | None = None,
    new_header: str = "",
    separator: str = "\n\n",
) -> dict:
    if not blocks:
        return {"text": "", "count": 0, "summary": "Порожній список блоків"}

    block_texts = []
    header_inserted = False

    for block in blocks:
        if not isinstance(block, dict):
            continue
        b_type = str(block.get("type", "")).casefold()
        b_text = str(block.get("text", "")).strip()

        if b_type == "header":
            if new_header.strip():
                block_texts.append(new_header.strip())
                header_inserted = True
            elif b_text:
                block_texts.append(b_text)
                header_inserted = True
        elif b_text:
            block_texts.append(b_text)

    # Якщо була вказана нова шапка, але блоку 'header' не було у списку — додаємо її на початок
    if new_header.strip() and not header_inserted:
        block_texts.insert(0, new_header.strip())

    sep = separator if separator else "\n\n"
    final_text = sep.join(block_texts).strip()

    summary = f"Зібрано наказ з {len(block_texts)} блоків"
    return {"text": final_text, "count": len(block_texts), "summary": summary}


@node(
    name="Розрахунок витягів з наказів",
    category="Наказ",
    description=(
        "Автоматично розраховує перелік витягів з наказу для кожної військової частини (ВЧ). "
        "Приймає наказ та таблицю частин, формує закриті коди частин (А1500, А1400) та виводить готовий список пунктів із дефісами (напр. А1500 1,3  А1400 1-6)."
    ),
    type_id="builtin.order.calculate_extracts",
    execution_inputs=("exec",),
    execution_outputs=("then",),
    outputs={
        "summary_text": "str",
        "table": "DataTable",
        "extracts": "dict",
        "count": "int",
        "summary": "str",
    },
)
def calculate_order_extracts(
    text: str = "",
    mapping: dict | None = None,
    header: str = "",
    rules: dict | list | str | None = None,
) -> dict:
    """Розпізнає частини у наказі та формує точний розрахунок витягів у форматі А1500 1,3 А1400 1-6."""
    if not text.strip():
        return {
            "summary_text": "",
            "table": DataTable(("Закрита ВЧ", "Відкрита назва", "Номери пунктів витягу", "Кількість пунктів"), ()),
            "extracts": {},
            "count": 0,
            "summary": "Порожній текст наказу",
        }

    body_text = text
    if rules:
        body_text, _ = _apply_custom_rules(body_text, rules)

    res = map_military_units(text=body_text, mapping=mapping)
    units_map = res.get("unit_paragraphs", {})
    abbr_map = res.get("unit_abbr_map", {})
    extracts_dict: dict[str, str] = {}
    table_rows = []
    summary_lines = []

    for sender, data in units_map.items():
        header_text = header
        if not header_text and isinstance(data, dict):
            header_text = "\n".join(data.get("header_lines", []))

        short_sender = sender
        open_names = ", ".join(sorted(res.get("unit_open_names", {}).get(sender, []))) if "unit_open_names" in res else ""

        items_text = ""
        range_labels = "-"
        count = 0

        if isinstance(data, dict) and "items" in data:
            section_groups: dict[str, list[str]] = {}
            raw_labels = []
            for item in data.get("items", []):
                heading = item.get("parent_heading", "").strip()
                item_text = item.get("text", "").strip()
                label = item.get("label", "")
                if label:
                    raw_labels.append(label)
                if item_text:
                    section_groups.setdefault(heading, []).append(item_text)

            range_labels = _format_item_numbers_range(raw_labels)
            count = len(data.get("items", []))

            body_sections = []
            for heading, item_texts in section_groups.items():
                if heading:
                    body_sections.append(f"{heading}\n\n" + "\n\n".join(item_texts))
                else:
                    body_sections.append("\n\n".join(item_texts))

            items_text = "\n\n".join(body_sections)
        elif isinstance(data, str):
            items_text = data
            count = 1

        full_doc = f"{header_text}\n\n{items_text}".strip() if header_text else items_text.strip()
        extracts_dict[short_sender] = full_doc
        table_rows.append((short_sender, open_names or "—", range_labels, count))
        summary_lines.append(f"{short_sender} {range_labels}")

    table = DataTable(
        ("Закрита ВЧ (Адресат)", "Відкрита назва ВЧ", "Номери пунктів витягу", "Кількість пунктів"),
        tuple(table_rows),
        "Розрахунок витягів з наказів",
    )
    summary_text = "\n".join(summary_lines)
    summary = f"Розраховано витягів: {len(extracts_dict)}"

    return {
        "summary_text": summary_text,
        "table": table,
        "extracts": extracts_dict,
        "count": len(extracts_dict),
        "summary": summary,
    }
