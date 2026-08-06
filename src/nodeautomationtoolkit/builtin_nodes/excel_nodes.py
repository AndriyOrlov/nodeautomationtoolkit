from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from nodeautomationtoolkit.core.definition import node
from nodeautomationtoolkit.core.table_types import DataTable


@node(
    name="Прочитати аркуш Excel",
    category="Excel",
    description="Зчитує дані з аркуша Excel (.xlsx) у вигляді структурованої таблиці DataTable.",
    type_id="builtin.excel.read_sheet",
    outputs={
        "table": "DataTable",
        "headers": "List",
        "rows_count": "int",
        "sheet_names": "List",
        "summary": "str",
    },
)
def read_sheet(
    excel_path: str = "",
    sheet_name: str = "",
    has_headers: bool = True,
    start_row: int = 1,
) -> dict:
    path = Path(excel_path).expanduser()
    if not path.is_file():
        raise FileNotFoundError(f"Excel-файл не знайдено: {excel_path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    all_rows = []
    for r in ws.iter_rows(min_row=max(1, start_row), values_only=True):
        row_vals = [str(val).strip() if val is not None else "" for val in r]
        if any(row_vals):
            all_rows.append(row_vals)

    if not all_rows:
        return {
            "table": DataTable((), ()),
            "headers": [],
            "rows_count": 0,
            "sheet_names": sheet_names,
            "summary": "Аркуш порожній",
        }

    if has_headers:
        headers = all_rows[0]
        data_rows = all_rows[1:]
    else:
        max_cols = max(len(r) for r in all_rows)
        headers = [f"Колонка {i+1}" for i in range(max_cols)]
        data_rows = all_rows

    col_count = len(headers)
    norm_rows = []
    for row in data_rows:
        padded = list(row[:col_count]) + [""] * max(0, col_count - len(row))
        norm_rows.append(tuple(padded))

    table = DataTable(tuple(headers), tuple(norm_rows), title=ws.title)
    summary = f"Прочитано {len(norm_rows)} рядків з аркуша '{ws.title}'"

    return {
        "table": table,
        "headers": headers,
        "rows_count": len(norm_rows),
        "sheet_names": sheet_names,
        "summary": summary,
    }


@node(
    name="Записати у комірку Excel",
    category="Excel",
    description="Записує або змінює значення у конкретній комірці Excel (наприклад, 'B5' або 'C12').",
    type_id="builtin.excel.write_cell",
    outputs={
        "output_path": "str",
        "summary": "str",
    },
    execution_inputs=("exec",),
    execution_outputs=("then",),
    preview_policy="never",
)
def write_cell(
    excel_path: str = "",
    cell_reference: str = "A1",
    value: Any = "",
    sheet_name: str = "",
    output_path: str = "",
) -> dict:
    src_path = Path(excel_path).expanduser()
    if not src_path.is_file():
        raise FileNotFoundError(f"Excel-файл не знайдено: {excel_path}")

    target_path = Path(output_path).expanduser() if output_path.strip() else src_path

    wb = openpyxl.load_workbook(src_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    ws[cell_reference.upper()] = value
    wb.save(target_path)

    summary = f"Значення '{value}' збережено у комірку {cell_reference.upper()} ({target_path.name})"
    return {"output_path": str(target_path), "summary": summary}


@node(
    name="Зберегти таблицю в Excel",
    category="Excel",
    description="Зберігає оброблену таблицю DataTable у новий чи існуючий файл Excel (.xlsx).",
    type_id="builtin.excel.save_table",
    outputs={
        "saved_path": "str",
        "rows_count": "int",
        "summary": "str",
    },
    execution_inputs=("exec",),
    execution_outputs=("then",),
    preview_policy="never",
)
def save_table(
    table: Any = None,
    output_path: str = "",
    sheet_name: str = "Результати",
) -> dict:
    if not output_path.strip():
        raise ValueError("Вкажіть шлях для збереження Excel-файла")

    out_path = Path(output_path).expanduser()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or "Аркуш1"

    rows_count = 0
    if isinstance(table, DataTable):
        ws.append(list(table.columns))
        for row in table.rows:
            ws.append(list(row))
            rows_count += 1
    elif isinstance(table, list):
        for item in table:
            if isinstance(item, (list, tuple)):
                ws.append(list(item))
                rows_count += 1
            else:
                ws.append([str(item)])
                rows_count += 1
    elif isinstance(table, dict):
        ws.append(["Ключ", "Значення"])
        for k, v in table.items():
            ws.append([str(k), str(v)])
            rows_count += 1

    wb.save(out_path)
    summary = f"Збережено {rows_count} рядків у {out_path.name}"
    return {"saved_path": str(out_path), "rows_count": rows_count, "summary": summary}


@node(
    name="Заміна тексту в Excel",
    category="Excel",
    description="Виконує масову заміну вказаного тексту у всіх комірках аркуша Excel.",
    type_id="builtin.excel.replace_text",
    outputs={
        "replaced_count": "int",
        "output_path": "str",
        "summary": "str",
    },
    execution_inputs=("exec",),
    execution_outputs=("then",),
    preview_policy="never",
)
def replace_text(
    excel_path: str = "",
    search_text: str = "",
    replace_text: str = "",
    sheet_name: str = "",
    output_path: str = "",
) -> dict:
    src_path = Path(excel_path).expanduser()
    if not src_path.is_file():
        raise FileNotFoundError(f"Excel-файл не знайдено: {excel_path}")

    target_path = Path(output_path).expanduser() if output_path.strip() else src_path
    wb = openpyxl.load_workbook(src_path)

    sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets

    replaced_count = 0
    for ws in sheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and search_text in cell.value:
                    cell.value = cell.value.replace(search_text, replace_text)
                    replaced_count += 1

    wb.save(target_path)
    summary = f"Виконано {replaced_count} замін у {target_path.name}"
    return {"replaced_count": replaced_count, "output_path": str(target_path), "summary": summary}
